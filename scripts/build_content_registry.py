"""
build_content_registry.py

Generates data/raw/ch04_content_registry.xlsx

Each sheet = one figure/table pair (or table-only).
Layout per sheet
----------------
Rows 1-N   : CONFIG block  (col A = key, col B = value)
Row N+1    : blank
Row N+2    : "DATA"  (marker so the loader knows where data starts)
Row N+3    : column headers
Row N+4+   : data rows

Config keys
-----------
chart_type          : line | clustered_bar | stacked_bar | simple_bar | horizontal_bar | table_only
figure_id           : Quarto label for the figure block (empty for table-only)
fig_label           : legacy alias for figure_id (kept for compatibility)
tbl_label           : quarto label for the table block
fig_cap             : figure caption   (empty for table-only)
tbl_cap             : table caption
x_col               : column to use as x-axis in the chart
x_axis_title        : x-axis label
y_axis_title        : y-axis label   (empty for table-only)
start_at_zero       : True | False
y_padding           : float (default 0.1)
hover_value_format  : e.g. .1f  .0f
hover_suffix        : e.g. %  (empty string if none)
value_format        : percent | number | currency | text  (drives table cell formatting)
pivot_for_chart     : True | False
  If True the raw data (x_col + county cols) is pivoted before charting:
  - the loader melts on x_col, pivots so the old x_col becomes series columns
    and "County" becomes the new x-axis column.
group_by            : x_col | series | (blank)
  Optional explicit grouping intent for bar charts:
  - x_col  -> counties on x-axis, grouped by x_col values (same effect as pivot_for_chart=True)
  - series -> keep x_col on x-axis, grouped by original series columns (same as pivot_for_chart=False)
  When provided, this overrides pivot_for_chart.
transpose_for_chart : True | False
  If True the data has category rows + county columns and is transposed
  so counties become the x-axis (used for poverty-by-race, alice-threshold,
  educational-attainment, modes-transportation).
y_cols_order        : pipe-separated list of y columns (in display order)
                      leave blank to use all non-x columns in sheet order
categorical_x       : True | False  (set x-axis to type="category")
has_multilevel_headers : True | False  (for tbl-poverty-threshold)
suppressed_marker   : value used for suppressed cells in data (default "s")
"""

import os
import math
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── output path ──────────────────────────────────────────────────────────────
OUT_DIR = os.path.join(os.path.dirname(__file__), "..", "data", "raw")
OUT_PATH = os.path.join(OUT_DIR, "ch04_content_registry.xlsx")

# ── style helpers ─────────────────────────────────────────────────────────────
HEADER_FILL   = PatternFill("solid", fgColor="1F4E79")   # dark blue
CONFIG_FILL   = PatternFill("solid", fgColor="D9E1F2")   # light blue
DATA_HDR_FILL = PatternFill("solid", fgColor="2E75B6")   # medium blue
ALT_ROW_FILL  = PatternFill("solid", fgColor="EBF3FB")   # very light blue
WHITE_FILL    = PatternFill("solid", fgColor="FFFFFF")

WHITE_FONT  = Font(bold=True, color="FFFFFF")
BOLD_FONT   = Font(bold=True)
NORMAL_FONT = Font()

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _write_sheet(wb: openpyxl.Workbook, sheet_name: str, config: dict, data: list[dict]):
    """
    Write one sheet.

    Parameters
    ----------
    config : dict   key-value pairs for the CONFIG block
    data   : list of dicts, each dict is one data row; keys = column headers
    """
    ws = wb.create_sheet(title=sheet_name)

    # ── CONFIG block ──────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 80

    # sheet title row
    title_cell = ws.cell(row=1, column=1, value=f"SHEET: {sheet_name}")
    title_cell.font = WHITE_FONT
    title_cell.fill = HEADER_FILL
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
    ws.row_dimensions[1].height = 20

    row = 2
    for key, val in config.items():
        cell_key = ws.cell(row=row, column=1, value=key)
        cell_val = ws.cell(row=row, column=2, value=str(val) if val is not None else "")
        cell_key.fill = CONFIG_FILL
        cell_val.fill = CONFIG_FILL
        cell_key.font = BOLD_FONT
        cell_key.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        cell_val.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        row += 1

    # blank separator
    row += 1

    # ── DATA block ────────────────────────────────────────────────────────────
    if not data:
        return

    columns = list(data[0].keys())

    # "DATA" marker
    marker = ws.cell(row=row, column=1, value="DATA")
    marker.font = WHITE_FONT
    marker.fill = HEADER_FILL
    marker.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(columns))
    ws.row_dimensions[row].height = 18
    row += 1

    # column headers
    for col_idx, col_name in enumerate(columns, start=1):
        c = ws.cell(row=row, column=col_idx, value=col_name)
        c.font = WHITE_FONT
        c.fill = DATA_HDR_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = max(
            14, len(str(col_name)) + 2
        )
    ws.row_dimensions[row].height = 30
    row += 1

    # data rows
    for r_idx, row_dict in enumerate(data):
        fill = ALT_ROW_FILL if r_idx % 2 == 1 else WHITE_FILL
        for col_idx, col_name in enumerate(columns, start=1):
            val = row_dict.get(col_name, "")
            # convert NaN → "s" (suppressed)
            if isinstance(val, float) and math.isnan(val):
                val = "s"
            c = ws.cell(row=row, column=col_idx, value=val)
            c.fill = fill
            c.border = THIN_BORDER
            c.alignment = Alignment(horizontal="center" if col_idx > 1 else "left",
                                    vertical="center", wrap_text=True)
        row += 1

    # freeze top rows (title + config + blank + marker + header)
    freeze_row = len(config) + 4   # 1 title + config rows + blank + marker + header
    ws.freeze_panes = ws.cell(row=freeze_row, column=1)


# ══════════════════════════════════════════════════════════════════════════════
#  SHEET DEFINITIONS
# ══════════════════════════════════════════════════════════════════════════════

def make_workbook() -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # remove default sheet

    # ── 1. labor-force ────────────────────────────────────────────────────────
    _write_sheet(wb, "labor-force", {
        "chart_type": "line",
        "fig_label": "fig-labor-force",
        "figure_id": "fig-labor-force",
        "tbl_label": "tbl-labor-force",
        "fig_cap": "Percentage of the Labor Force, Population 16 Years and Older, 2021-2023",
        "tbl_cap": "Percentage of the Labor Force, Population 16 Years and Older, 2021-2023",
        "x_col": "Year",
        "x_axis_title": "Year",
        "y_axis_title": "Percentage of Labor Force (%)",
        "start_at_zero": "False",
        "y_padding": "0.1",
        "hover_value_format": ".1f",
        "hover_suffix": "%",
        "value_format": "percent",
        "pivot_for_chart": "False",
        "transpose_for_chart": "False",
        "y_cols_order": "",
        "categorical_x": "False",
        "has_multilevel_headers": "False",
    }, [
        {"Year": 2021, "Dutchess": 63.0, "Orange": 63.9, "Putnam": 65.2, "Rockland": 63.9,
         "Sullivan": 56.1, "Ulster": 60.1, "Westchester": 65.5, "NYS": 63.1, "US": 63.6},
        {"Year": 2022, "Dutchess": 62.8, "Orange": 63.4, "Putnam": 64.7, "Rockland": 63.2,
         "Sullivan": 58.1, "Ulster": 58.9, "Westchester": 65.2, "NYS": 62.9, "US": 63.5},
        {"Year": 2023, "Dutchess": 63.3, "Orange": 63.5, "Putnam": 64.7, "Rockland": 63.2,
         "Sullivan": 59.1, "Ulster": 58.7, "Westchester": 65.4, "NYS": 63.0, "US": 63.5},
    ])

    # ── 2. unemployment ───────────────────────────────────────────────────────
    _write_sheet(wb, "unemployment", {
        "chart_type": "clustered_bar",
        "fig_label": "fig-unemployment",
        "figure_id": "fig-unemployment",
        "tbl_label": "tbl-unemployment",
        "fig_cap": "Percentage of Population Unemployed, 16 Years and Older, 2021-2023",
        "tbl_cap": "Percentage of Population Unemployed, 16 Years and Older, 2021-2023",
        "x_col": "Year",
        "x_axis_title": "County",
        "y_axis_title": "Percentage Unemployed (%)",
        "start_at_zero": "False",
        "y_padding": "0.1",
        "hover_value_format": ".1f",
        "hover_suffix": "%",
        "value_format": "percent",
        "pivot_for_chart": "True",
        "transpose_for_chart": "False",
        "y_cols_order": "",
        "categorical_x": "True",
        "has_multilevel_headers": "False",
    }, [
        {"Year": 2021, "Dutchess": 5.2, "Orange": 5.4, "Putnam": 4.8, "Rockland": 6.0,
         "Sullivan": 8.0, "Ulster": 5.1, "Westchester": 6.1, "NYS": 6.2, "US": 5.5},
        {"Year": 2022, "Dutchess": 5.0, "Orange": 5.3, "Putnam": 4.4, "Rockland": 6.2,
         "Sullivan": 7.2, "Ulster": 5.1, "Westchester": 6.0, "NYS": 6.2, "US": 5.3},
        {"Year": 2023, "Dutchess": 4.8, "Orange": 5.4, "Putnam": 4.1, "Rockland": 5.9,
         "Sullivan": 6.1, "Ulster": 5.1, "Westchester": 6.0, "NYS": 6.2, "US": 5.2},
    ])

    # ── 3. food-insecurity ────────────────────────────────────────────────────
    _write_sheet(wb, "food-insecurity", {
        "chart_type": "line",
        "fig_label": "fig-food-insecurity",
        "figure_id": "fig-food-insecurity",
        "tbl_label": "tbl-food-insecurity",
        "fig_cap": "Percentage of Overall Food Insecurity, 2020-2023",
        "tbl_cap": "Percentage of Overall Food Insecurity, 2020-2023",
        "x_col": "Year",
        "x_axis_title": "Year",
        "y_axis_title": "Food Insecurity Rate (%)",
        "start_at_zero": "False",
        "y_padding": "0.1",
        "hover_value_format": ".1f",
        "hover_suffix": "%",
        "value_format": "percent",
        "pivot_for_chart": "False",
        "transpose_for_chart": "False",
        "y_cols_order": "",
        "categorical_x": "False",
        "has_multilevel_headers": "False",
    }, [
        {"Year": 2020, "Dutchess": 8.7, "Orange": 9.4, "Putnam": 6.3, "Rockland": 9.7,
         "Sullivan": 11.5, "Ulster": 11.3, "Westchester": 7.9, "NYS": 9.6},
        {"Year": 2021, "Dutchess": 7.2, "Orange": 7.8, "Putnam": 5.3, "Rockland": 8.2,
         "Sullivan": 9.9, "Ulster": 9.5, "Westchester": 6.6, "NYS": 11.4},
        {"Year": 2022, "Dutchess": 10.0, "Orange": 11.3, "Putnam": 8.4, "Rockland": 11.0,
         "Sullivan": 13.1, "Ulster": 12.8, "Westchester": 9.4, "NYS": 13.4},
        {"Year": 2023, "Dutchess": 10.8, "Orange": 12.1, "Putnam": 8.9, "Rockland": 12.0,
         "Sullivan": 14.0, "Ulster": 13.2, "Westchester": 10.7, "NYS": 14.5},
    ])

    # ── 4. child-food-insecurity ───────────────────────────────────────────────
    _write_sheet(wb, "child-food-insecurity", {
        "chart_type": "line",
        "fig_label": "fig-childhood-food-insecurity",
        "figure_id": "fig-childhood-food-insecurity",
        "tbl_label": "tbl-childhood-food-insecurity",
        "fig_cap": "Percentage of Food Insecurity, Children 18 Years and Younger, 2020-2023",
        "tbl_cap": "Percentage of Food Insecurity, Children 18 Years and Younger, 2020-2023",
        "x_col": "Year",
        "x_axis_title": "Year",
        "y_axis_title": "Child Food Insecurity Rate (%)",
        "start_at_zero": "False",
        "y_padding": "0.1",
        "hover_value_format": ".1f",
        "hover_suffix": "%",
        "value_format": "percent",
        "pivot_for_chart": "False",
        "transpose_for_chart": "False",
        "y_cols_order": "",
        "categorical_x": "False",
        "has_multilevel_headers": "False",
    }, [
        {"Year": 2020, "Dutchess": 11.5, "Orange": 13.0, "Putnam": 7.8, "Rockland": 13.5,
         "Sullivan": 16.5, "Ulster": 15.8, "Westchester": 10.2, "NYS": 14.0},
        {"Year": 2021, "Dutchess": 9.8, "Orange": 11.0, "Putnam": 6.5, "Rockland": 11.5,
         "Sullivan": 14.2, "Ulster": 13.5, "Westchester": 8.9, "NYS": 16.0},
        {"Year": 2022, "Dutchess": 13.5, "Orange": 15.8, "Putnam": 9.8, "Rockland": 15.5,
         "Sullivan": 18.5, "Ulster": 17.8, "Westchester": 12.8, "NYS": 19.0},
        {"Year": 2023, "Dutchess": 14.8, "Orange": 16.8, "Putnam": 6.6, "Rockland": 16.5,
         "Sullivan": 19.9, "Ulster": 18.5, "Westchester": 14.0, "NYS": 20.5},
    ])

    # ── 5. cost-burdened-renters ───────────────────────────────────────────────
    _write_sheet(wb, "cost-burdened-renters", {
        "chart_type": "clustered_bar",
        "fig_label": "fig-cost-burdened-renters",
        "figure_id": "fig-cost-burdened-renters",
        "tbl_label": "tbl-cost-burdened-renters",
        "fig_cap": "Percentage of Cost Burdened Renter Occupied Units, 2021-2023",
        "tbl_cap": "Cost Burdened Renter Occupied Units, 2021-2023",
        "x_col": "Year",
        "x_axis_title": "County",
        "y_axis_title": "Percentage of Cost Burdened Renter Occupied Units (%)",
        "start_at_zero": "False",
        "y_padding": "0.1",
        "hover_value_format": ".1f",
        "hover_suffix": "%",
        "value_format": "percent",
        "pivot_for_chart": "True",
        "transpose_for_chart": "False",
        "y_cols_order": "",
        "categorical_x": "True",
        "has_multilevel_headers": "False",
    }, [
        {"Year": 2021, "Dutchess": 52.3, "Orange": 56.5, "Putnam": 52.7, "Rockland": 58.3,
         "Sullivan": 48.0, "Ulster": 55.3, "Westchester": 53.2, "NYS": 51.6},
        {"Year": 2022, "Dutchess": 52.4, "Orange": 56.1, "Putnam": 53.2, "Rockland": 58.9,
         "Sullivan": 48.4, "Ulster": 55.3, "Westchester": 53.5, "NYS": 51.7},
        {"Year": 2023, "Dutchess": 52.0, "Orange": 56.2, "Putnam": 56.5, "Rockland": 59.6,
         "Sullivan": 48.7, "Ulster": 56.6, "Westchester": 53.0, "NYS": 51.5},
    ])

    # ── 6. severely-cost-burdened ──────────────────────────────────────────────
    _write_sheet(wb, "severely-cost-burdened", {
        "chart_type": "clustered_bar",
        "fig_label": "fig-severely-cost-burdened",
        "figure_id": "fig-severely-cost-burdened",
        "tbl_label": "tbl-severely-cost-burdened",
        "fig_cap": "Percentage of Severely Cost Burdened Households, 2016-2023",
        "tbl_cap": "Severely Cost Burdened Households, 2016-2020 to 2019-2023",
        "x_col": "Period",
        "x_axis_title": "County",
        "y_axis_title": "Percentage of Severely Cost Burdened Households (%)",
        "start_at_zero": "False",
        "y_padding": "0.1",
        "hover_value_format": ".1f",
        "hover_suffix": "%",
        "value_format": "percent",
        "pivot_for_chart": "True",
        "transpose_for_chart": "False",
        "y_cols_order": "",
        "categorical_x": "True",
        "has_multilevel_headers": "False",
    }, [
        {"Period": "2016–2020", "Dutchess": 16.0, "Orange": 19.0, "Putnam": 16.0,
         "Rockland": 22.0, "Sullivan": 14.0, "Ulster": 18.0, "Westchester": 20.0, "NYS": 19.0},
        {"Period": "2017–2021", "Dutchess": 16.0, "Orange": 19.0, "Putnam": 16.0,
         "Rockland": 22.0, "Sullivan": 15.0, "Ulster": 18.0, "Westchester": 19.0, "NYS": 19.0},
        {"Period": "2018–2022", "Dutchess": 16.0, "Orange": 20.0, "Putnam": 16.0,
         "Rockland": 22.0, "Sullivan": 15.0, "Ulster": 18.0, "Westchester": 20.0, "NYS": 19.0},
        {"Period": "2019–2023", "Dutchess": 16.0, "Orange": 20.0, "Putnam": 17.0,
         "Rockland": 22.0, "Sullivan": 15.0, "Ulster": 18.0, "Westchester": 19.0, "NYS": 19.0},
    ])

    # ── 7. hud-housing ────────────────────────────────────────────────────────
    _write_sheet(wb, "hud-housing", {
        "chart_type": "clustered_bar",
        "fig_label": "fig-hud-housing",
        "figure_id": "fig-hud-housing",
        "tbl_label": "tbl-hud-housing",
        "fig_cap": "Number of People Living in HUD-subsidized Housing in the Past 12 Months, 2021–2024",
        "tbl_cap": "Number of People Living in Housing and Urban Development (HUD)-subsidized Housing in the Past 12 Months, 2021–2024",
        "x_col": "Year",
        "x_axis_title": "County",
        "y_axis_title": "Number of People",
        "start_at_zero": "True",
        "y_padding": "0.1",
        "hover_value_format": ".0f",
        "hover_suffix": "",
        "value_format": "number",
        "pivot_for_chart": "True",
        "transpose_for_chart": "False",
        "y_cols_order": "",
        "categorical_x": "True",
        "has_multilevel_headers": "False",
    }, [
        {"Year": 2021, "Dutchess": 7442, "Orange": 18258, "Putnam": 945, "Rockland": 21732,
         "Sullivan": 5018, "Ulster": 5484, "Westchester": 40230, "NYS": 1025652},
        {"Year": 2022, "Dutchess": 7641, "Orange": 18745, "Putnam": 929, "Rockland": 22170,
         "Sullivan": 5228, "Ulster": 5479, "Westchester": 40412, "NYS": 985104},
        {"Year": 2023, "Dutchess": 7630, "Orange": 19000, "Putnam": 992, "Rockland": 23411,
         "Sullivan": 4846, "Ulster": 5418, "Westchester": 40415, "NYS": 987957},
        {"Year": 2024, "Dutchess": 7484, "Orange": 19129, "Putnam": 1012, "Rockland": 23735,
         "Sullivan": 4445, "Ulster": 4980, "Westchester": 40137, "NYS": 1000730},
    ])

    # ── 8. poverty-threshold  (TABLE ONLY – multilevel headers) ───────────────
    # Flatten the multilevel structure for Excel (two header rows)
    _write_sheet(wb, "poverty-threshold", {
        "chart_type": "table_only",
        "fig_label": "",
        "tbl_label": "tbl-poverty-threshold",
        "fig_cap": "",
        "tbl_cap": "Poverty Threshold for 2024 by Size of Family and Number of Related Children 18 Years and Younger",
        "x_col": "Family Size",
        "x_axis_title": "",
        "y_axis_title": "",
        "start_at_zero": "",
        "y_padding": "",
        "hover_value_format": "",
        "hover_suffix": "",
        "value_format": "currency",
        "pivot_for_chart": "False",
        "transpose_for_chart": "False",
        "y_cols_order": "",
        "categorical_x": "False",
        "has_multilevel_headers": "True",
        "note": "Two header rows: row 1 = top-level group label (Related children under 18 years), row 2 = sub-level (None, One, Two, …). The DATA section stores only the values; reconstruct headers in code.",
    }, [
        {"Family Size": "One person (unrelated individual): Under age 65",
         "None": "$16,320", "One": "", "Two": "", "Three": "", "Four": "",
         "Five": "", "Six": "", "Seven": "", "Eight": "", "Nine or more": ""},
        {"Family Size": "One person (unrelated individual): Aged 65 and older",
         "None": "$15,045", "One": "", "Two": "", "Three": "", "Four": "",
         "Five": "", "Six": "", "Seven": "", "Eight": "", "Nine or more": ""},
        {"Family Size": "Two people: Householder under age 65",
         "None": "$21,006", "One": "$21,621", "Two": "", "Three": "", "Four": "",
         "Five": "", "Six": "", "Seven": "", "Eight": "", "Nine or more": ""},
        {"Family Size": "Two people: Householder aged 65 and older",
         "None": "$18,961", "One": "$21,540", "Two": "", "Three": "", "Four": "",
         "Five": "", "Six": "", "Seven": "", "Eight": "", "Nine or more": ""},
        {"Family Size": "Three people",
         "None": "$24,537", "One": "$25,249", "Two": "$25,273", "Three": "", "Four": "",
         "Five": "", "Six": "", "Seven": "", "Eight": "", "Nine or more": ""},
        {"Family Size": "Four people",
         "None": "$32,355", "One": "$32,884", "Two": "$31,812", "Three": "$31,922", "Four": "",
         "Five": "", "Six": "", "Seven": "", "Eight": "", "Nine or more": ""},
        {"Family Size": "Five people",
         "None": "$39,019", "One": "$39,586", "Two": "$38,374", "Three": "$37,436",
         "Four": "$36,863", "Five": "", "Six": "", "Seven": "", "Eight": "", "Nine or more": ""},
        {"Family Size": "Six people",
         "None": "$44,879", "One": "$45,057", "Two": "$44,128", "Three": "$43,238",
         "Four": "$41,915", "Five": "$41,131", "Six": "", "Seven": "", "Eight": "", "Nine or more": ""},
        {"Family Size": "Seven people",
         "None": "$51,638", "One": "$51,961", "Two": "$50,849", "Three": "$50,075",
         "Four": "$48,631", "Five": "$46,948", "Six": "$45,100", "Seven": "", "Eight": "", "Nine or more": ""},
        {"Family Size": "Eight people",
         "None": "$57,753", "One": "$58,263", "Two": "$57,215", "Three": "$56,296",
         "Four": "$54,992", "Five": "$53,337", "Six": "$51,614", "Seven": "$51,177",
         "Eight": "", "Nine or more": ""},
        {"Family Size": "Nine people or more",
         "None": "$69,473", "One": "$69,810", "Two": "$68,882", "Three": "$68,102",
         "Four": "$66,822", "Five": "$65,062", "Six": "$63,469", "Seven": "$63,075",
         "Eight": "$60,645", "Nine or more": "$60,645"},
    ])

    # ── 9. poverty-rate ───────────────────────────────────────────────────────
    _write_sheet(wb, "poverty-rate", {
        "chart_type": "clustered_bar",
        "fig_label": "fig-poverty-rate",
        "figure_id": "fig-poverty-rate",
        "tbl_label": "tbl-poverty-rate",
        "fig_cap": "Percentage of Population in Poverty, 2021-2023",
        "tbl_cap": "Percentage of Population in Poverty, 2021-2023",
        "x_col": "Year",
        "x_axis_title": "County",
        "y_axis_title": "Poverty Rate (%)",
        "start_at_zero": "True",
        "y_padding": "0.1",
        "hover_value_format": ".1f",
        "hover_suffix": "%",
        "value_format": "percent",
        "pivot_for_chart": "True",
        "transpose_for_chart": "False",
        "y_cols_order": "",
        "categorical_x": "True",
        "has_multilevel_headers": "False",
    }, [
        {"Year": 2021, "Dutchess": 8.8, "Orange": 11.7, "Putnam": 6.0, "Rockland": 14.9,
         "Sullivan": 14.1, "Ulster": 13.2, "Westchester": 8.2, "NYS": 13.5},
        {"Year": 2022, "Dutchess": 8.6, "Orange": 13.0, "Putnam": 6.3, "Rockland": 15.1,
         "Sullivan": 14.8, "Ulster": 14.7, "Westchester": 8.5, "NYS": 13.6},
        {"Year": 2023, "Dutchess": 8.3, "Orange": 13.0, "Putnam": 6.5, "Rockland": 15.6,
         "Sullivan": 15.2, "Ulster": 14.3, "Westchester": 8.9, "NYS": 13.7},
    ])

    # ── 10. poverty-by-race ────────────────────────────────────────────────────
    _write_sheet(wb, "poverty-by-race", {
        "chart_type": "clustered_bar",
        "fig_label": "fig-poverty-by-race",
        "figure_id": "fig-poverty-by-race",
        "tbl_label": "tbl-poverty-by-race",
        "fig_cap": "Percentage of Families Below Poverty Level by Race and Ethnicity, 2023",
        "tbl_cap": "Percentage of Families Below Poverty Level by Race and Ethnicity, 2023",
        "x_col": "Race/Ethnicity",
        "x_axis_title": "County",
        "y_axis_title": "Poverty Rate (%)",
        "start_at_zero": "True",
        "y_padding": "0.1",
        "hover_value_format": ".1f",
        "hover_suffix": "%",
        "value_format": "percent",
        "pivot_for_chart": "False",
        "transpose_for_chart": "True",
        "y_cols_order": "White (non-Hispanic)|Black (including Hispanic)|Asian (including Hispanic, excluding PI)|Hispanic (any race)|Total",
        "categorical_x": "True",
        "has_multilevel_headers": "False",
    }, [
        {"Race/Ethnicity": "White (non-Hispanic)", "Dutchess": 3.3, "Orange": 7.9, "Putnam": 2.8,
         "Rockland": 11.1, "Sullivan": 8.2, "Ulster": 4.9, "Westchester": 2.6, "NYS": 5.7},
        {"Race/Ethnicity": "Black (including Hispanic)", "Dutchess": 14.2, "Orange": 12.5,
         "Putnam": 0.0, "Rockland": 8.6, "Sullivan": 30.5, "Ulster": 14.1, "Westchester": 10.8, "NYS": 16.5},
        {"Race/Ethnicity": "Asian (including Hispanic, excluding PI)", "Dutchess": 7.5,
         "Orange": 10.3, "Putnam": 23.4, "Rockland": 1.8, "Sullivan": 13.3, "Ulster": 4.8,
         "Westchester": 5.4, "NYS": 11.1},
        {"Race/Ethnicity": "Hispanic (any race)", "Dutchess": 11.2, "Orange": 11.3,
         "Putnam": 6.6, "Rockland": 12.8, "Sullivan": 14.3, "Ulster": 24.4,
         "Westchester": 10.4, "NYS": 17.1},
        {"Race/Ethnicity": "Total", "Dutchess": 5.3, "Orange": 9.2, "Putnam": 3.8,
         "Rockland": 10.3, "Sullivan": 10.5, "Ulster": 7.1, "Westchester": 5.7, "NYS": 9.7},
    ])

    # ── 11. economically-disadvantaged ────────────────────────────────────────
    _write_sheet(wb, "economically-disadvantaged", {
        "chart_type": "line",
        "fig_label": "fig-economically-disadvantaged",
        "figure_id": "fig-economically-disadvantaged",
        "tbl_label": "tbl-economically-disadvantaged",
        "fig_cap": "Enrollment Rate of Economically Disadvantaged Students, 2021-2024",
        "tbl_cap": "Enrollment Rate of Economically Disadvantaged Students, 2021-2024",
        "x_col": "School Year",
        "x_axis_title": "School Year",
        "y_axis_title": "Economically Disadvantaged Students (%)",
        "start_at_zero": "False",
        "y_padding": "0.1",
        "hover_value_format": ".1f",
        "hover_suffix": "%",
        "value_format": "percent",
        "pivot_for_chart": "False",
        "transpose_for_chart": "False",
        "y_cols_order": "",
        "categorical_x": "False",
        "has_multilevel_headers": "False",
    }, [
        {"School Year": "2021-2022", "Dutchess": 39.2, "Orange": 47.0, "Putnam": 28.5,
         "Rockland": 43.9, "Sullivan": 60.4, "Ulster": 48.6, "Westchester": 38.9, "NYS": 56.2},
        {"School Year": "2022-2023", "Dutchess": 43.1, "Orange": 49.0, "Putnam": 32.1,
         "Rockland": 48.1, "Sullivan": 62.7, "Ulster": 52.5, "Westchester": 39.6, "NYS": 59.1},
        {"School Year": "2023-2024", "Dutchess": 43.3, "Orange": 49.5, "Putnam": 32.5,
         "Rockland": 51.0, "Sullivan": 61.4, "Ulster": 48.8, "Westchester": 39.7, "NYS": 59.2},
    ])

    # ── 12. alice-budget  (TABLE ONLY) ────────────────────────────────────────
    _write_sheet(wb, "alice-budget", {
        "chart_type": "table_only",
        "fig_label": "",
        "tbl_label": "tbl-alice-budget",
        "fig_cap": "",
        "tbl_cap": "ALICE Household Survival Budget, New York State, 2023",
        "x_col": "Budget Item",
        "x_axis_title": "",
        "y_axis_title": "",
        "start_at_zero": "",
        "y_padding": "",
        "hover_value_format": "",
        "hover_suffix": "",
        "value_format": "currency",
        "pivot_for_chart": "False",
        "transpose_for_chart": "False",
        "y_cols_order": "",
        "categorical_x": "False",
        "has_multilevel_headers": "False",
    }, [
        {"Budget Item": "Housing", "Single Adult": 1103, "One Adult, One Child": 1189,
         "One Adult, One in Child Care": 1189, "Two Adults": 1189,
         "Two Adults, Two Children": 1437, "Two Adults, Two in Child Care": 1437,
         "Single Adult 65+": 1103, "Two Adults 65+": 1189},
        {"Budget Item": "Child Care", "Single Adult": 0, "One Adult, One Child": 423,
         "One Adult, One in Child Care": 1126, "Two Adults": 0,
         "Two Adults, Two Children": 844, "Two Adults, Two in Child Care": 2345,
         "Single Adult 65+": 0, "Two Adults 65+": 0},
        {"Budget Item": "Food", "Single Adult": 516, "One Adult, One Child": 873,
         "One Adult, One in Child Care": 783, "Two Adults": 946,
         "Two Adults, Two Children": 1587, "Two Adults, Two in Child Care": 1400,
         "Single Adult 65+": 475, "Two Adults 65+": 870},
        {"Budget Item": "Transportation", "Single Adult": 401, "One Adult, One Child": 535,
         "One Adult, One in Child Care": 506, "Two Adults": 617,
         "Two Adults, Two Children": 957, "Two Adults, Two in Child Care": 899,
         "Single Adult 65+": 346, "Two Adults 65+": 508},
        {"Budget Item": "Health Care", "Single Adult": 196, "One Adult, One Child": 452,
         "One Adult, One in Child Care": 452, "Two Adults": 452,
         "Two Adults, Two Children": 775, "Two Adults, Two in Child Care": 775,
         "Single Adult 65+": 543, "Two Adults 65+": 1086},
        {"Budget Item": "Technology", "Single Adult": 86, "One Adult, One Child": 86,
         "One Adult, One in Child Care": 86, "Two Adults": 116,
         "Two Adults, Two Children": 116, "Two Adults, Two in Child Care": 116,
         "Single Adult 65+": 86, "Two Adults 65+": 116},
        {"Budget Item": "Miscellaneous", "Single Adult": 230, "One Adult, One Child": 356,
         "One Adult, One in Child Care": 414, "Two Adults": 332,
         "Two Adults, Two Children": 572, "Two Adults, Two in Child Care": 697,
         "Single Adult 65+": 255, "Two Adults 65+": 377},
        {"Budget Item": "Taxes", "Single Adult": 439, "One Adult, One Child": 461,
         "One Adult, One in Child Care": 625, "Two Adults": 529,
         "Two Adults, Two Children": 684, "Two Adults, Two in Child Care": 1037,
         "Single Adult 65+": 510, "Two Adults 65+": 861},
        {"Budget Item": "Monthly Total", "Single Adult": 2971, "One Adult, One Child": 4375,
         "One Adult, One in Child Care": 5181, "Two Adults": 4181,
         "Two Adults, Two Children": 6972, "Two Adults, Two in Child Care": 8706,
         "Single Adult 65+": 3318, "Two Adults 65+": 5007},
        {"Budget Item": "ANNUAL TOTAL", "Single Adult": 35652, "One Adult, One Child": 52500,
         "One Adult, One in Child Care": 62172, "Two Adults": 50172,
         "Two Adults, Two Children": 83664, "Two Adults, Two in Child Care": 104472,
         "Single Adult 65+": 39816, "Two Adults 65+": 60084},
        {"Budget Item": "Hourly Wage", "Single Adult": 17.83, "One Adult, One Child": 26.25,
         "One Adult, One in Child Care": 31.09, "Two Adults": 25.09,
         "Two Adults, Two Children": 41.83, "Two Adults, Two in Child Care": 52.24,
         "Single Adult 65+": 19.91, "Two Adults 65+": 30.04},
    ])

    # ── 13. alice-threshold ────────────────────────────────────────────────────
    _write_sheet(wb, "alice-threshold", {
        "chart_type": "stacked_bar",
        "fig_label": "fig-alice-threshold",
        "figure_id": "fig-alice-threshold",
        "tbl_label": "tbl-alice-threshold",
        "fig_cap": "ALICE Threshold Percentage, 2023",
        "tbl_cap": "ALICE Threshold Percentage, 2023",
        "x_col": "Category",
        "x_axis_title": "County",
        "y_axis_title": "Percentage (%)",
        "start_at_zero": "True",
        "y_padding": "0.1",
        "hover_value_format": ".1f",
        "hover_suffix": "%",
        "value_format": "percent",
        "pivot_for_chart": "False",
        "transpose_for_chart": "True",
        "y_cols_order": "ALICE|Above ALICE Threshold|Poverty",
        "categorical_x": "True",
        "has_multilevel_headers": "False",
    }, [
        {"Category": "ALICE", "Dutchess": 31.0, "Orange": 33.0, "Putnam": 32.0,
         "Rockland": 38.0, "Sullivan": 30.0, "Ulster": 32.0, "Westchester": 28.0, "NYS": 33.0},
        {"Category": "Poverty", "Dutchess": 6.0, "Orange": 10.0, "Putnam": 6.0,
         "Rockland": 12.0, "Sullivan": 15.0, "Ulster": 10.0, "Westchester": 10.0, "NYS": 14.0},
        {"Category": "Above ALICE Threshold", "Dutchess": 62.0, "Orange": 57.0, "Putnam": 63.0,
         "Rockland": 50.0, "Sullivan": 54.0, "Ulster": 58.0, "Westchester": 61.0, "NYS": 52.0},
    ])

    # ── 14. graduation-rate ────────────────────────────────────────────────────
    _write_sheet(wb, "graduation-rate", {
        "chart_type": "clustered_bar",
        "fig_label": "fig-graduation-rate",
        "figure_id": "fig-graduation-rate",
        "tbl_label": "tbl-graduation-rate",
        "fig_cap": "High School Graduation Rate, 2021-2023",
        "tbl_cap": "High School Graduation Rate, 2021-2023",
        "x_col": "Year",
        "x_axis_title": "County",
        "y_axis_title": "High School Graduation Rate (%)",
        "start_at_zero": "False",
        "y_padding": "0.1",
        "hover_value_format": ".1f",
        "hover_suffix": "%",
        "value_format": "percent",
        "pivot_for_chart": "True",
        "transpose_for_chart": "False",
        "y_cols_order": "",
        "categorical_x": "True",
        "has_multilevel_headers": "False",
    }, [
        {"Year": 2021, "Dutchess": 87.0, "Orange": 89.0, "Putnam": 94.0, "Rockland": 90.0,
         "Sullivan": 82.0, "Ulster": 87.0, "Westchester": 91.0, "NYS": 86.0},
        {"Year": 2022, "Dutchess": 86.0, "Orange": 89.0, "Putnam": 94.0, "Rockland": 88.0,
         "Sullivan": 78.0, "Ulster": 87.0, "Westchester": 92.0, "NYS": 87.0},
        {"Year": 2023, "Dutchess": 87.0, "Orange": 89.0, "Putnam": 91.0, "Rockland": 86.0,
         "Sullivan": 76.0, "Ulster": 87.0, "Westchester": 91.0, "NYS": 86.0},
    ])

    # ── 15. graduation-by-race ─────────────────────────────────────────────────
    _write_sheet(wb, "graduation-by-race", {
        "chart_type": "clustered_bar",
        "fig_label": "fig-graduation-by-race",
        "figure_id": "fig-graduation-by-race",
        "tbl_label": "tbl-graduation-by-race",
        "fig_cap": "High School Graduation Rate, by Race and Ethnicity, 2023",
        "tbl_cap": "High School Graduation Rate, by Race and Ethnicity, 2023",
        "x_col": "Race/Ethnicity",
        "x_axis_title": "Race/Ethnicity",
        "y_axis_title": "High School Graduation Rate (%)",
        "start_at_zero": "True",
        "y_padding": "0.1",
        "hover_value_format": ".1f",
        "hover_suffix": "%",
        "value_format": "percent",
        "pivot_for_chart": "False",
        "transpose_for_chart": "False",
        "y_cols_order": "",
        "categorical_x": "True",
        "has_multilevel_headers": "False",
        "suppressed_marker": "s",
    }, [
        {"Race/Ethnicity": "Asian or Native Hawaiian/Other PI",
         "Dutchess": 91.0, "Orange": 96.0, "Putnam": 100.0, "Rockland": 97.0,
         "Sullivan": float("nan"), "Ulster": float("nan"), "Westchester": 97.0, "NYS": 93.0},
        {"Race/Ethnicity": "Black or African American",
         "Dutchess": 78.0, "Orange": 85.0, "Putnam": 97.0, "Rockland": 86.0,
         "Sullivan": 66.0, "Ulster": 84.0, "Westchester": 83.0, "NYS": 81.0},
        {"Race/Ethnicity": "White",
         "Dutchess": 92.0, "Orange": 93.0, "Putnam": 95.0, "Rockland": 94.0,
         "Sullivan": 84.0, "Ulster": 90.0, "Westchester": 96.0, "NYS": 91.0},
        {"Race/Ethnicity": "Multiracial",
         "Dutchess": 80.0, "Orange": 88.0, "Putnam": float("nan"), "Rockland": float("nan"),
         "Sullivan": 68.0, "Ulster": 74.0, "Westchester": 94.0, "NYS": 84.0},
        {"Race/Ethnicity": "Hispanic",
         "Dutchess": 79.0, "Orange": 83.0, "Putnam": 80.0, "Rockland": 74.0,
         "Sullivan": 63.0, "Ulster": 80.0, "Westchester": 85.0, "NYS": 81.0},
    ])

    # ── 16. educational-attainment ─────────────────────────────────────────────
    _write_sheet(wb, "educational-attainment", {
        "chart_type": "stacked_bar",
        "fig_label": "fig-educational-attainment",
        "figure_id": "fig-educational-attainment",
        "tbl_label": "tbl-educational-attainment",
        "fig_cap": "Rate of Education Attainment, 2023",
        "tbl_cap": "Rate of Education Attainment, 2023",
        "x_col": "Education Level",
        "x_axis_title": "Education Level",
        "y_axis_title": "Percentage (%)",
        "start_at_zero": "True",
        "y_padding": "0.1",
        "hover_value_format": ".1f",
        "hover_suffix": "%",
        "value_format": "percent",
        "pivot_for_chart": "False",
        "transpose_for_chart": "False",
        "y_cols_order": "",
        "categorical_x": "True",
        "has_multilevel_headers": "False",
    }, [
        {"Education Level": "Some college, no degree",
         "Dutchess": 16.9, "Orange": 19.1, "Putnam": 16.5, "Rockland": 16.2,
         "Sullivan": 17.9, "Ulster": 17.1, "Westchester": 12.6, "NYS": 14.9},
        {"Education Level": "Associate's degree",
         "Dutchess": 10.1, "Orange": 10.3, "Putnam": 8.4, "Rockland": 7.8,
         "Sullivan": 10.3, "Ulster": 9.6, "Westchester": 6.4, "NYS": 8.9},
        {"Education Level": "Bachelor's degree",
         "Dutchess": 21.0, "Orange": 18.2, "Putnam": 24.2, "Rockland": 23.2,
         "Sullivan": 17.0, "Ulster": 20.2, "Westchester": 26.0, "NYS": 22.0},
        {"Education Level": "Graduate or professional degree",
         "Dutchess": 18.7, "Orange": 14.1, "Putnam": 19.3, "Rockland": 18.9,
         "Sullivan": 12.7, "Ulster": 16.9, "Westchester": 26.5, "NYS": 17.5},
    ])

    # ── 17. language-proficiency ───────────────────────────────────────────────
    _write_sheet(wb, "language-proficiency", {
        "chart_type": "clustered_bar",
        "fig_label": "fig-language-proficiency",
        "figure_id": "fig-language-proficiency",
        "tbl_label": "tbl-language-proficiency",
        "fig_cap": "Percentage of Population that Speaks English Less Than Very Well, 5 Years and Older, 2021-2023",
        "tbl_cap": "Percentage of Population that Speaks English Less Than Very Well, 5 Years and Older, 2021-2023",
        "x_col": "Year",
        "x_axis_title": "County",
        "y_axis_title": "Percentage Speaking English Less Than Very Well (%)",
        "start_at_zero": "False",
        "y_padding": "0.1",
        "hover_value_format": ".1f",
        "hover_suffix": "%",
        "value_format": "percent",
        "pivot_for_chart": "True",
        "transpose_for_chart": "False",
        "y_cols_order": "",
        "categorical_x": "True",
        "has_multilevel_headers": "False",
    }, [
        {"Year": 2021, "Dutchess": 4.7, "Orange": 10.6, "Putnam": 5.3, "Rockland": 18.6,
         "Sullivan": 5.9, "Ulster": 2.9, "Westchester": 12.7, "NYS": 13.1},
        {"Year": 2022, "Dutchess": 4.8, "Orange": 11.3, "Putnam": 5.6, "Rockland": 18.8,
         "Sullivan": 7.3, "Ulster": 3.0, "Westchester": 12.2, "NYS": 13.1},
        {"Year": 2023, "Dutchess": 5.4, "Orange": 12.2, "Putnam": 6.0, "Rockland": 20.0,
         "Sullivan": 8.0, "Ulster": 3.4, "Westchester": 12.4, "NYS": 13.3},
    ])

    # ── 18. disconnected-youth ─────────────────────────────────────────────────
    _write_sheet(wb, "disconnected-youth", {
        "chart_type": "clustered_bar",
        "fig_label": "fig-disconnected-youth",
        "figure_id": "fig-disconnected-youth",
        "tbl_label": "tbl-disconnected-youth",
        "fig_cap": "Percentage of Disconnected Youth Ages 16-19, 2014-2023",
        "tbl_cap": "Percentage of Disconnected Youth Ages 16-19, 2014-2023",
        "x_col": "Period",
        "x_axis_title": "County",
        "y_axis_title": "Percentage of Disconnected Youth (%)",
        "start_at_zero": "False",
        "y_padding": "0.1",
        "hover_value_format": ".1f",
        "hover_suffix": "%",
        "value_format": "percent",
        "pivot_for_chart": "True",
        "transpose_for_chart": "False",
        "y_cols_order": "",
        "categorical_x": "True",
        "has_multilevel_headers": "False",
        "suppressed_marker": "s",
    }, [
        {"Period": "2014–2018", "Dutchess": 4.0, "Orange": 8.0, "Putnam": 4.0,
         "Rockland": 5.0, "Sullivan": 12.0, "Ulster": 6.0, "Westchester": 6.0, "NYS": 6.0},
        {"Period": "2015–2019", "Dutchess": 5.0, "Orange": 8.0, "Putnam": float("nan"),
         "Rockland": 5.0, "Sullivan": 12.0, "Ulster": 6.0, "Westchester": 6.0, "NYS": 6.0},
        {"Period": "2016–2020", "Dutchess": 5.0, "Orange": 10.0, "Putnam": float("nan"),
         "Rockland": 5.0, "Sullivan": 17.0, "Ulster": 6.0, "Westchester": 4.0, "NYS": 6.0},
        {"Period": "2017–2021", "Dutchess": 6.0, "Orange": 8.0, "Putnam": float("nan"),
         "Rockland": 4.0, "Sullivan": 17.0, "Ulster": 6.0, "Westchester": 4.0, "NYS": 6.0},
        {"Period": "2018–2022", "Dutchess": 7.0, "Orange": 8.0, "Putnam": float("nan"),
         "Rockland": 6.0, "Sullivan": 25.0, "Ulster": 9.0, "Westchester": 5.0, "NYS": 7.0},
        {"Period": "2019–2023", "Dutchess": 6.0, "Orange": 8.0, "Putnam": float("nan"),
         "Rockland": 6.0, "Sullivan": 24.0, "Ulster": 9.0, "Westchester": 5.0, "NYS": 7.0},
    ])

    # ── 19. residential-segregation  (TABLE ONLY) ──────────────────────────────
    _write_sheet(wb, "residential-segregation", {
        "chart_type": "table_only",
        "fig_label": "",
        "tbl_label": "tbl-residential-segregation",
        "fig_cap": "",
        "tbl_cap": "Index Score of Residential Segregation, 2013-2023",
        "x_col": "Period",
        "x_axis_title": "",
        "y_axis_title": "",
        "start_at_zero": "",
        "y_padding": "",
        "hover_value_format": "",
        "hover_suffix": "",
        "value_format": "number",
        "pivot_for_chart": "False",
        "transpose_for_chart": "False",
        "y_cols_order": "",
        "categorical_x": "False",
        "has_multilevel_headers": "False",
        "suppressed_marker": "s",
    }, [
        {"Period": "2013–2017", "Dutchess": 52.0, "Orange": 44.0, "Putnam": 39.0,
         "Rockland": 58.0, "Sullivan": 46.0, "Ulster": 49.0, "Westchester": 62.0, "NYS": 74.0},
        {"Period": "2016–2020", "Dutchess": 50.0, "Orange": 45.0, "Putnam": 44.0,
         "Rockland": 55.0, "Sullivan": 50.0, "Ulster": 50.0, "Westchester": 59.0, "NYS": 74.0},
        {"Period": "2017–2021", "Dutchess": 47.0, "Orange": 47.0, "Putnam": 38.0,
         "Rockland": 58.0, "Sullivan": 55.0, "Ulster": 46.0, "Westchester": 59.0, "NYS": 74.0},
        {"Period": "2018–2022", "Dutchess": 46.0, "Orange": 49.0, "Putnam": 45.0,
         "Rockland": 56.0, "Sullivan": 51.0, "Ulster": 43.0, "Westchester": 60.0, "NYS": 74.0},
        {"Period": "2019–2023", "Dutchess": 49.0, "Orange": 48.0, "Putnam": 49.0,
         "Rockland": 58.0, "Sullivan": 43.0, "Ulster": float("nan"),
         "Westchester": 63.0, "NYS": 75.0},
    ])

    # ── 20. food-environment-index ─────────────────────────────────────────────
    _write_sheet(wb, "food-environment-index", {
        "chart_type": "clustered_bar",
        "fig_label": "fig-food-environment-index",
        "figure_id": "fig-food-environment-index",
        "tbl_label": "tbl-food-environment-index",
        "fig_cap": "Food Environment Index, 2025",
        "tbl_cap": "Food Environment Index, 2025",
        "x_col": "County",
        "x_axis_title": "County",
        "y_axis_title": "Food Environment Index (0-10)",
        "start_at_zero": "False",
        "y_padding": "0.1",
        "hover_value_format": ".1f",
        "hover_suffix": "",
        "value_format": "number",
        "pivot_for_chart": "False",
        "transpose_for_chart": "False",
        "y_cols_order": "Food Environment Index",
        "categorical_x": "True",
        "has_multilevel_headers": "False",
    }, [
        {"County": "Dutchess", "Food Environment Index": 8.4},
        {"County": "Orange", "Food Environment Index": 8.3},
        {"County": "Putnam", "Food Environment Index": 9.0},
        {"County": "Rockland", "Food Environment Index": 8.5},
        {"County": "Sullivan", "Food Environment Index": 8.2},
        {"County": "Ulster", "Food Environment Index": 8.1},
        {"County": "Westchester", "Food Environment Index": 9.3},
        {"County": "NYS", "Food Environment Index": 8.7},
    ])

    # ── 21. limited-access-foods ───────────────────────────────────────────────
    _write_sheet(wb, "limited-access-foods", {
        "chart_type": "clustered_bar",
        "fig_label": "fig-limited-access-healthy-foods",
        "figure_id": "fig-limited-access-healthy-foods",
        "tbl_label": "tbl-limited-access-healthy-foods",
        "fig_cap": "Percentage of Population with Limited Access to Healthy Foods, 2025",
        "tbl_cap": "Percentage of Population with Limited Access to Healthy Foods, 2025",
        "x_col": "County",
        "x_axis_title": "County",
        "y_axis_title": "Percentage with Limited Access (%)",
        "start_at_zero": "False",
        "y_padding": "0.1",
        "hover_value_format": ".1f",
        "hover_suffix": "%",
        "value_format": "percent",
        "pivot_for_chart": "False",
        "transpose_for_chart": "False",
        "y_cols_order": "Limited Access",
        "categorical_x": "True",
        "has_multilevel_headers": "False",
    }, [
        {"County": "Dutchess", "Limited Access": 2.1},
        {"County": "Orange", "Limited Access": 2.8},
        {"County": "Putnam", "Limited Access": 6.7},
        {"County": "Rockland", "Limited Access": 2.5},
        {"County": "Sullivan", "Limited Access": 3.2},
        {"County": "Ulster", "Limited Access": 2.3},
        {"County": "Westchester", "Limited Access": 1.4},
        {"County": "NYS", "Limited Access": 2.0},
    ])

    # ── 22. violent-crime ──────────────────────────────────────────────────────
    _write_sheet(wb, "violent-crime", {
        "chart_type": "clustered_bar",
        "fig_label": "fig-violent-crime",
        "figure_id": "fig-violent-crime",
        "tbl_label": "tbl-violent-crime",
        "fig_cap": "Violent Crime Rate per 100,000 Population, 2018-2021",
        "tbl_cap": "Violent Crime Rate per 100,000 Population, 2018-2021",
        "x_col": "Year",
        "x_axis_title": "County",
        "y_axis_title": "Violent Crime Rate per 100,000",
        "start_at_zero": "False",
        "y_padding": "0.1",
        "hover_value_format": ".1f",
        "hover_suffix": "",
        "value_format": "number",
        "pivot_for_chart": "True",
        "transpose_for_chart": "False",
        "y_cols_order": "",
        "categorical_x": "True",
        "has_multilevel_headers": "False",
    }, [
        {"Year": 2018, "Dutchess": 118.0, "Orange": 185.0, "Putnam": 45.0, "Rockland": 95.0,
         "Sullivan": 165.0, "Ulster": 125.0, "Westchester": 110.0, "NYS": 380.0},
        {"Year": 2019, "Dutchess": 120.0, "Orange": 190.0, "Putnam": 48.0, "Rockland": 98.0,
         "Sullivan": 170.0, "Ulster": 128.0, "Westchester": 112.0, "NYS": 375.0},
        {"Year": 2020, "Dutchess": 115.0, "Orange": 188.0, "Putnam": 42.0, "Rockland": 92.0,
         "Sullivan": 155.0, "Ulster": 120.0, "Westchester": 108.0, "NYS": 370.0},
        {"Year": 2021, "Dutchess": 112.0, "Orange": 192.4, "Putnam": 40.0, "Rockland": 88.0,
         "Sullivan": 145.0, "Ulster": 118.0, "Westchester": 105.0, "NYS": 365.0},
    ])

    # ── 23. air-pollution ──────────────────────────────────────────────────────
    _write_sheet(wb, "air-pollution", {
        "chart_type": "line",
        "fig_label": "fig-air-pollution",
        "figure_id": "fig-air-pollution",
        "tbl_label": "tbl-air-pollution",
        "fig_cap": "Average Daily Density of Fine Particulate Matter, 2014-2020",
        "tbl_cap": "Average Daily Density of Fine Particulate Matter, 2014-2020",
        "x_col": "Year",
        "x_axis_title": "Year",
        "y_axis_title": "Fine Particulate Matter (μg/m³)",
        "start_at_zero": "False",
        "y_padding": "0.1",
        "hover_value_format": ".1f",
        "hover_suffix": " μg/m³",
        "value_format": "number",
        "pivot_for_chart": "False",
        "transpose_for_chart": "False",
        "y_cols_order": "",
        "categorical_x": "False",
        "has_multilevel_headers": "False",
    }, [
        {"Year": 2014, "Dutchess": 8.5, "Orange": 7.8, "Putnam": 8.2, "Rockland": 8.8,
         "Sullivan": 7.5, "Ulster": 8.0, "Westchester": 9.2, "NYS": 8.5},
        {"Year": 2015, "Dutchess": 8.2, "Orange": 7.5, "Putnam": 8.0, "Rockland": 8.5,
         "Sullivan": 7.2, "Ulster": 7.8, "Westchester": 9.0, "NYS": 8.3},
        {"Year": 2016, "Dutchess": 7.9, "Orange": 7.2, "Putnam": 7.8, "Rockland": 8.2,
         "Sullivan": 7.0, "Ulster": 7.5, "Westchester": 8.8, "NYS": 8.1},
        {"Year": 2017, "Dutchess": 7.6, "Orange": 7.0, "Putnam": 7.5, "Rockland": 8.0,
         "Sullivan": 6.8, "Ulster": 7.3, "Westchester": 8.6, "NYS": 7.9},
        {"Year": 2018, "Dutchess": 7.4, "Orange": 6.8, "Putnam": 7.3, "Rockland": 7.8,
         "Sullivan": 6.6, "Ulster": 7.1, "Westchester": 8.4, "NYS": 7.7},
        {"Year": 2019, "Dutchess": 7.2, "Orange": 6.6, "Putnam": 7.1, "Rockland": 7.6,
         "Sullivan": 6.4, "Ulster": 6.9, "Westchester": 8.2, "NYS": 7.5},
        {"Year": 2020, "Dutchess": 7.0, "Orange": 6.4, "Putnam": 6.9, "Rockland": 7.4,
         "Sullivan": 6.2, "Ulster": 6.7, "Westchester": 8.0, "NYS": 7.3},
    ])

    # ── 24. lead-testing ───────────────────────────────────────────────────────
    _write_sheet(wb, "lead-testing", {
        "chart_type": "clustered_bar",
        "fig_label": "fig-lead-testing",
        "figure_id": "fig-lead-testing",
        "tbl_label": "tbl-lead-testing",
        "fig_cap": "Percentage of Children Tested for Lead at Least Twice Before 36 Months of Age, 2016-2019 Birth Cohorts",
        "tbl_cap": "Percentage of Children Tested for Lead at Least Twice Before 36 Months of Age, 2016-2019 Birth Cohorts",
        "x_col": "Birth Cohort",
        "x_axis_title": "County",
        "y_axis_title": "Percentage Tested (%)",
        "start_at_zero": "False",
        "y_padding": "0.1",
        "hover_value_format": ".1f",
        "hover_suffix": "%",
        "value_format": "percent",
        "pivot_for_chart": "True",
        "transpose_for_chart": "False",
        "y_cols_order": "",
        "categorical_x": "True",
        "has_multilevel_headers": "False",
    }, [
        {"Birth Cohort": 2016, "Dutchess": 45.2, "Orange": 42.3, "Putnam": 48.5,
         "Rockland": 44.8, "Sullivan": 32.5, "Ulster": 43.2, "Westchester": 56.8, "NYS": 52.5},
        {"Birth Cohort": 2017, "Dutchess": 46.8, "Orange": 43.7, "Putnam": 50.2,
         "Rockland": 46.2, "Sullivan": 34.1, "Ulster": 44.6, "Westchester": 58.2, "NYS": 53.8},
        {"Birth Cohort": 2018, "Dutchess": 48.1, "Orange": 45.2, "Putnam": 51.8,
         "Rockland": 47.5, "Sullivan": 35.2, "Ulster": 45.9, "Westchester": 59.5, "NYS": 55.2},
        {"Birth Cohort": 2019, "Dutchess": 49.5, "Orange": 46.8, "Putnam": 53.2,
         "Rockland": 48.9, "Sullivan": 36.6, "Ulster": 47.2, "Westchester": 60.2, "NYS": 56.5},
    ])

    # ── 25. severe-housing-problems ────────────────────────────────────────────
    _write_sheet(wb, "severe-housing-problems", {
        "chart_type": "clustered_bar",
        "fig_label": "fig-severe-housing-problems",
        "figure_id": "fig-severe-housing-problems",
        "tbl_label": "tbl-severe-housing-problems",
        "fig_cap": "Percentage of Households with Severe Housing Problems, 2025",
        "tbl_cap": "Percentage of Households with Severe Housing Problems, 2025",
        "x_col": "County",
        "x_axis_title": "County",
        "y_axis_title": "Percentage with Severe Housing Problems (%)",
        "start_at_zero": "False",
        "y_padding": "0.1",
        "hover_value_format": ".1f",
        "hover_suffix": "%",
        "value_format": "percent",
        "pivot_for_chart": "False",
        "transpose_for_chart": "False",
        "y_cols_order": "Severe Housing Problems",
        "categorical_x": "True",
        "has_multilevel_headers": "False",
    }, [
        {"County": "Dutchess", "Severe Housing Problems": 20.0},
        {"County": "Orange", "Severe Housing Problems": 22.0},
        {"County": "Putnam", "Severe Housing Problems": 17.0},
        {"County": "Rockland", "Severe Housing Problems": 26.0},
        {"County": "Sullivan", "Severe Housing Problems": 16.0},
        {"County": "Ulster", "Severe Housing Problems": 19.0},
        {"County": "Westchester", "Severe Housing Problems": 21.0},
        {"County": "NYS", "Severe Housing Problems": 23.0},
    ])

    # ── 26. no-vehicles ────────────────────────────────────────────────────────
    _write_sheet(wb, "no-vehicles", {
        "chart_type": "clustered_bar",
        "fig_label": "fig-no-vehicles",
        "figure_id": "fig-no-vehicles",
        "tbl_label": "tbl-no-vehicles",
        "fig_cap": "Percentage of Households with No Available Vehicles, 2023",
        "tbl_cap": "Percentage of Households with No Available Vehicles, 2023",
        "x_col": "County",
        "x_axis_title": "County",
        "y_axis_title": "Percentage with No Vehicles (%)",
        "start_at_zero": "False",
        "y_padding": "0.1",
        "hover_value_format": ".1f",
        "hover_suffix": "%",
        "value_format": "percent",
        "pivot_for_chart": "False",
        "transpose_for_chart": "False",
        "y_cols_order": "No Vehicles",
        "categorical_x": "True",
        "has_multilevel_headers": "False",
    }, [
        {"County": "Dutchess", "No Vehicles": 6.8},
        {"County": "Orange", "No Vehicles": 8.2},
        {"County": "Putnam", "No Vehicles": 4.2},
        {"County": "Rockland", "No Vehicles": 9.5},
        {"County": "Sullivan", "No Vehicles": 7.2},
        {"County": "Ulster", "No Vehicles": 6.5},
        {"County": "Westchester", "No Vehicles": 14.2},
        {"County": "NYS", "No Vehicles": 10.5},
    ])

    # ── 27. modes-transportation ───────────────────────────────────────────────
    _write_sheet(wb, "modes-transportation", {
        "chart_type": "stacked_bar",
        "fig_label": "fig-modes-transportation",
        "figure_id": "fig-modes-transportation",
        "tbl_label": "tbl-modes-transportation",
        "fig_cap": "Modes of Transportation to Work, 2023",
        "tbl_cap": "Modes of Transportation to Work, 2023",
        "x_col": "Mode",
        "x_axis_title": "Mode of Transportation",
        "y_axis_title": "Percentage (%)",
        "start_at_zero": "True",
        "y_padding": "0.1",
        "hover_value_format": ".1f",
        "hover_suffix": "%",
        "value_format": "percent",
        "pivot_for_chart": "False",
        "transpose_for_chart": "False",
        "y_cols_order": "",
        "categorical_x": "True",
        "has_multilevel_headers": "False",
    }, [
        {"Mode": "Drive alone", "Dutchess": 68.5, "Orange": 70.2, "Putnam": 65.8,
         "Rockland": 66.5, "Sullivan": 74.1, "Ulster": 63.2, "Westchester": 51.9, "NYS": 58.5},
        {"Mode": "Carpool", "Dutchess": 8.2, "Orange": 7.8, "Putnam": 9.5,
         "Rockland": 8.8, "Sullivan": 6.5, "Ulster": 7.5, "Westchester": 6.8, "NYS": 7.2},
        {"Mode": "Public transportation", "Dutchess": 3.5, "Orange": 2.5, "Putnam": 4.2,
         "Rockland": 8.5, "Sullivan": 1.2, "Ulster": 2.8, "Westchester": 17.8, "NYS": 12.5},
        {"Mode": "Walk", "Dutchess": 2.8, "Orange": 2.2, "Putnam": 2.5,
         "Rockland": 2.8, "Sullivan": 1.8, "Ulster": 3.2, "Westchester": 3.5, "NYS": 3.8},
        {"Mode": "Other", "Dutchess": 1.2, "Orange": 1.1, "Putnam": 1.3,
         "Rockland": 1.2, "Sullivan": 0.9, "Ulster": 1.0, "Westchester": 1.2, "NYS": 1.5},
        {"Mode": "Work from home", "Dutchess": 12.8, "Orange": 11.2, "Putnam": 13.7,
         "Rockland": 8.2, "Sullivan": 8.5, "Ulster": 15.3, "Westchester": 17.3, "NYS": 12.5},
    ])

    # ── 28. commute-time ──────────────────────────────────────────────────────
    _write_sheet(wb, "commute-time", {
        "chart_type": "clustered_bar",
        "fig_label": "fig-commute-time",
        "figure_id": "fig-commute-time",
        "tbl_label": "tbl-commute-time",
        "fig_cap": "Average Commute Time to Work, 2023",
        "tbl_cap": "Average Commute Time to Work, 2023",
        "x_col": "County",
        "x_axis_title": "County",
        "y_axis_title": "Average Commute Time (minutes)",
        "start_at_zero": "False",
        "y_padding": "0.1",
        "hover_value_format": ".0f",
        "hover_suffix": " minutes",
        "value_format": "number",
        "pivot_for_chart": "False",
        "transpose_for_chart": "False",
        "y_cols_order": "Average Commute Time (minutes)",
        "categorical_x": "True",
        "has_multilevel_headers": "False",
    }, [
        {"County": "Dutchess", "Average Commute Time (minutes)": 28},
        {"County": "Orange", "Average Commute Time (minutes)": 34},
        {"County": "Putnam", "Average Commute Time (minutes)": 39},
        {"County": "Rockland", "Average Commute Time (minutes)": 32},
        {"County": "Sullivan", "Average Commute Time (minutes)": 25},
        {"County": "Ulster", "Average Commute Time (minutes)": 25},
        {"County": "Westchester", "Average Commute Time (minutes)": 35},
        {"County": "NYS", "Average Commute Time (minutes)": 33},
    ])

    return wb


# ── main ──────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    os.makedirs(OUT_DIR, exist_ok=True)
    wb = make_workbook()
    wb.save(OUT_PATH)
    print(f"✓  Saved {OUT_PATH}")
    print(f"   Sheets: {wb.sheetnames}")
