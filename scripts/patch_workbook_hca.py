"""
patch_workbook_hca.py

Second-pass workbook patch that:

  1. Recreates the Residential-Segregation sheet as type="both" (adding a
     Line figure so fig-residential-segregation works in the QMD).
  2. Adds 8 Health Care Access indicators extracted from the ch04 QMD
     hardcoded blocks:
       - medical-care-cost
       - children-uninsured
       - adults-insured
       - medically-underserved  (table only)
       - primary-care-providers
       - dentists
       - mental-health-providers
       - access-to-primary-care

Run from the project root after migrate_workbook.py:
    python scripts/patch_workbook_hca.py
"""

from __future__ import annotations

import math
import sys
from pathlib import Path

PROJECT_ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from openpyxl import load_workbook

WORKBOOK_PATH = PROJECT_ROOT / "data" / "raw" / "Mid-Hudson Regional Community Health Assessment 2025 Data File.xlsx"

# Re-use helpers from migrate_workbook
from scripts.migrate_workbook import _write_flat_sheet

# ── indicator definitions ─────────────────────────────────────────────────────

PATCH_INDICATORS: list[dict] = [
    # ── Residential Segregation (upgrade to both) ─────────────────────────────
    # This replaces the table-only sheet created in migrate_workbook.py
    {
        "sheet_name": "Residential-Segregation",
        "object_id": "residential-segregation",
        "caption": "Index Score of Residential Segregation, 2013-2023",
        "type": "both",
        "figure_type": "Line",
        "x_col": "Period",
        "y_cols": "",
        "y_axis_title": "Index of Dissimilarity",
        "start_at_zero": False,
        "pivot_for_chart": False,
        "data_type": "Number",
        "multilevel_headers": False,
        "source": {
            "Table ID": "",
            "URL": "https://www.countyhealthrankings.org/",
            "Data Year": 2023,
            "Estimate Type": "",
            "Citation Month": "April",
            "Citation Year": 2025,
            "Custom Text": "County Health Rankings and Roadmaps, 2025",
        },
        "data": [
            {"Period": "2013\u20132017", "Dutchess": 52.0, "Orange": 44.0, "Putnam": 39.0,
             "Rockland": 58.0, "Sullivan": 46.0, "Ulster": 49.0, "Westchester": 62.0, "NYS": 74.0},
            {"Period": "2016\u20132020", "Dutchess": 50.0, "Orange": 45.0, "Putnam": 44.0,
             "Rockland": 55.0, "Sullivan": 50.0, "Ulster": 50.0, "Westchester": 59.0, "NYS": 74.0},
            {"Period": "2017\u20132021", "Dutchess": 47.0, "Orange": 47.0, "Putnam": 38.0,
             "Rockland": 58.0, "Sullivan": 55.0, "Ulster": 46.0, "Westchester": 59.0, "NYS": 74.0},
            {"Period": "2018\u20132022", "Dutchess": 46.0, "Orange": 49.0, "Putnam": 45.0,
             "Rockland": 56.0, "Sullivan": 51.0, "Ulster": 43.0, "Westchester": 60.0, "NYS": 74.0},
            {"Period": "2019\u20132023", "Dutchess": 49.0, "Orange": 48.0, "Putnam": 49.0,
             "Rockland": 58.0, "Sullivan": 43.0, "Ulster": "s",
             "Westchester": 63.0, "NYS": 75.0},
        ],
    },
    # ── Medical Care Cost ─────────────────────────────────────────────────────
    {
        "sheet_name": "Medical-Care-Cost",
        "object_id": "medical-care-cost",
        "caption": "Percentage of Adults Who Did Not Receive Medical Care Due to Cost, 2016, 2018, 2021",
        "type": "both",
        "figure_type": "Line",
        "x_col": "Year",
        "y_cols": "",
        "y_axis_title": "Percentage of Adults (%)",
        "start_at_zero": False,
        "pivot_for_chart": False,
        "data_type": "Percent",
        "multilevel_headers": False,
        "source": {
            "Table ID": "",
            "URL": "https://webbi1.health.ny.gov/SASStoredProcess/guest?_program=%2FEBI%2FPHIG%2Fapps%2Fchir_dashboard%2Fchir_dashboard",
            "Data Year": 2021,
            "Estimate Type": "",
            "Citation Month": "April",
            "Citation Year": 2025,
            "Custom Text": "NYS Community Health Indicator Reports Dashboard, April 2025",
        },
        "data": [
            {"Year": 2016, "Dutchess": 8.6, "Orange": 11.1, "Putnam": 13.5, "Rockland": 11.8,
             "Sullivan": 20.9, "Ulster": 11.2, "Westchester": 12.4, "Mid-Hudson": 11.8, "NYS": 11.5},
            {"Year": 2018, "Dutchess": 7.7, "Orange": 8.5, "Putnam": 11.4, "Rockland": 11.8,
             "Sullivan": 11.3, "Ulster": 12.7, "Westchester": 7.5, "Mid-Hudson": 8.4, "NYS": 11.3},
            {"Year": 2021, "Dutchess": 8.7, "Orange": "5.0*", "Putnam": 9.9, "Rockland": 8.4,
             "Sullivan": 4.9, "Ulster": 5.3, "Westchester": 8.3, "Mid-Hudson": 7.4, "NYS": 8.4},
        ],
    },
    # ── Children Without Health Insurance ────────────────────────────────────
    {
        "sheet_name": "Children-Uninsured",
        "object_id": "children-uninsured",
        "caption": "Percentage of Children Without Health Insurance, 2020-2023",
        "type": "both",
        "figure_type": "Line",
        "x_col": "Year",
        "y_cols": "",
        "y_axis_title": "Percentage of Children Without Health Insurance (%)",
        "start_at_zero": False,
        "pivot_for_chart": False,
        "data_type": "Percent",
        "multilevel_headers": False,
        "source": {
            "Table ID": "S2701",
            "URL": "https://data.census.gov/table/ACSST5Y2023.S2701",
            "Data Year": 2023,
            "Estimate Type": "5-Year Estimates",
            "Citation Month": "April",
            "Citation Year": 2025,
            "Custom Text": "",
        },
        "data": [
            {"Year": 2020, "Dutchess": 2.4, "Orange": 2.3, "Putnam": 2.5, "Rockland": 2.4,
             "Sullivan": 4.1, "Ulster": 2.5, "Westchester": 2.7, "NYS": 2.5},
            {"Year": 2021, "Dutchess": 2.7, "Orange": 2.8, "Putnam": 2.5, "Rockland": 2.3,
             "Sullivan": 3.2, "Ulster": 2.8, "Westchester": 3.2, "NYS": 2.6},
            {"Year": 2022, "Dutchess": 2.3, "Orange": 2.8, "Putnam": 2.5, "Rockland": 2.3,
             "Sullivan": 3.4, "Ulster": 2.7, "Westchester": 2.3, "NYS": 2.5},
            {"Year": 2023, "Dutchess": 2.6, "Orange": 2.7, "Putnam": 2.4, "Rockland": 2.8,
             "Sullivan": 3.5, "Ulster": 3.0, "Westchester": 2.8, "NYS": 2.8},
        ],
    },
    # ── Adults with Health Insurance ──────────────────────────────────────────
    {
        "sheet_name": "Adults-Insured",
        "object_id": "adults-insured",
        "caption": "Percentage of Adults with Health Insurance, 2020-2023",
        "type": "both",
        "figure_type": "Line",
        "x_col": "Year",
        "y_cols": "",
        "y_axis_title": "Percentage of Adults with Health Insurance (%)",
        "start_at_zero": False,
        "pivot_for_chart": False,
        "data_type": "Percent",
        "multilevel_headers": False,
        "source": {
            "Table ID": "S2701",
            "URL": "https://data.census.gov/table/ACSST5Y2023.S2701",
            "Data Year": 2023,
            "Estimate Type": "5-Year Estimates",
            "Citation Month": "April",
            "Citation Year": 2025,
            "Custom Text": "",
        },
        "data": [
            {"Year": 2020, "Dutchess": 94.5, "Orange": 93.2, "Putnam": 95.1, "Rockland": 93.3,
             "Sullivan": 92.6, "Ulster": 92.6, "Westchester": 93.2, "NYS": 92.7},
            {"Year": 2021, "Dutchess": 94.2, "Orange": 93.5, "Putnam": 94.6, "Rockland": 93.3,
             "Sullivan": 92.4, "Ulster": 92.7, "Westchester": 93.0, "NYS": 92.6},
            {"Year": 2022, "Dutchess": 94.6, "Orange": 93.3, "Putnam": 95.0, "Rockland": 93.6,
             "Sullivan": 92.9, "Ulster": 91.8, "Westchester": 93.6, "NYS": 93.2},
            {"Year": 2023, "Dutchess": 94.4, "Orange": 93.0, "Putnam": 94.9, "Rockland": 92.3,
             "Sullivan": 92.5, "Ulster": 92.0, "Westchester": 93.6, "NYS": 93.2},
        ],
    },
    # ── Medically Underserved (table only) ───────────────────────────────────
    {
        "sheet_name": "Medically-Underserved",
        "object_id": "medically-underserved",
        "caption": "Medically Underserved Areas and Medically Underserved Populations",
        "type": "table",
        "figure_type": "",
        "x_col": "County",
        "y_cols": "",
        "y_axis_title": "",
        "start_at_zero": False,
        "pivot_for_chart": False,
        "data_type": "Number",
        "multilevel_headers": False,
        "source": {
            "Table ID": "",
            "URL": "https://data.hrsa.gov/tools/shortage-area",
            "Data Year": 2025,
            "Estimate Type": "",
            "Citation Month": "April",
            "Citation Year": 2025,
            "Custom Text": "Health Resources and Services Administration, April 2025",
        },
        "data": [
            {"County": "Dutchess", "Area Name": "Low Income - Poughkeepsie",
             "Designation Type": "MUP Low Income", "IMU Score": 59.2, "Designation Date": "5/25/2001"},
            {"County": "Dutchess", "Area Name": "Migrant & Seasonal Farm Worker - East Dutchess",
             "Designation Type": "MUP Low Income", "IMU Score": 44.8, "Designation Date": "4/16/2001"},
            {"County": "Dutchess", "Area Name": "Medicaid Eligible and Medically Indigent - Beacon Service Area",
             "Designation Type": "MUP Other Population Governor's Exception", "IMU Score": "s", "Designation Date": "7/6/1993"},
            {"County": "Orange", "Area Name": "Orange County Service Area",
             "Designation Type": "Medically Underserved Area", "IMU Score": 55.5, "Designation Date": "5/4/1994"},
            {"County": "Orange", "Area Name": "Village of Walden Service Area",
             "Designation Type": "Medically Underserved Area", "IMU Score": 60.8, "Designation Date": "6/29/1999"},
            {"County": "Orange", "Area Name": "Village of Kiryas Joel Service Area",
             "Designation Type": "Medically Underserved Area", "IMU Score": 45.0, "Designation Date": "7/21/1993"},
            {"County": "Orange", "Area Name": "Low Income - Middletown Service Area",
             "Designation Type": "MUP Low Income", "IMU Score": 58.2, "Designation Date": "4/8/1994"},
            {"County": "Rockland", "Area Name": "Village of New Square Service Area",
             "Designation Type": "Medically Underserved Area", "IMU Score": 45.5, "Designation Date": "8/3/1993"},
            {"County": "Rockland", "Area Name": "Low Income - Haverstraw",
             "Designation Type": "MUP Low Income", "IMU Score": 61.6, "Designation Date": "7/27/2006"},
            {"County": "Sullivan", "Area Name": "Low Income - Western Sullivan Service Area",
             "Designation Type": "MUP Low Income", "IMU Score": 59.3, "Designation Date": "5/31/2002"},
            {"County": "Sullivan", "Area Name": "Low Income - Monticello",
             "Designation Type": "MUP Low Income", "IMU Score": 61.4, "Designation Date": "6/24/2004"},
            {"County": "Sullivan and Ulster", "Area Name": "Low Income - Wawarsing/Fallsburg S Area",
             "Designation Type": "MUP Low Income", "IMU Score": 61.8, "Designation Date": "6/18/2002"},
            {"County": "Ulster", "Area Name": "Plattekill Town - County",
             "Designation Type": "Medically Underserved Area", "IMU Score": 58.8, "Designation Date": "5/7/1981"},
            {"County": "Westchester", "Area Name": "Westchester Service Area - Elmsford",
             "Designation Type": "Medically Underserved Area", "IMU Score": 61.6, "Designation Date": "7/5/1994"},
            {"County": "Westchester", "Area Name": "Westchester Service Area - Mount Vernon",
             "Designation Type": "Medically Underserved Area", "IMU Score": 54.0, "Designation Date": "4/6/1978"},
            {"County": "Westchester", "Area Name": "Low Income - Mount Kisco",
             "Designation Type": "MUP Other Population Governor's Exception", "IMU Score": "s", "Designation Date": "2/28/2003"},
            {"County": "Westchester", "Area Name": "Westchester Service Area - Peekskill",
             "Designation Type": "Medically Underserved Area", "IMU Score": 58.8, "Designation Date": "5/4/1994"},
            {"County": "Westchester", "Area Name": "Medicaid Eligible and Medically Indigent - Port Chester",
             "Designation Type": "MUP Other Population Governor's Exception", "IMU Score": "s", "Designation Date": "4/8/1993"},
            {"County": "Westchester", "Area Name": "Westchester Service Area - Yonkers",
             "Designation Type": "Medically Underserved Area", "IMU Score": 41.2, "Designation Date": "10/7/1988"},
        ],
    },
    # ── Primary Care Providers ────────────────────────────────────────────────
    {
        "sheet_name": "Primary-Care-Providers",
        "object_id": "primary-care-providers",
        "caption": "Ratio of Residents to Primary Care Providers, 2019-2021",
        "type": "both",
        "figure_type": "Line",
        "x_col": "Year",
        "y_cols": "",
        "y_axis_title": "Ratio of Residents to Primary Care Providers",
        "start_at_zero": False,
        "pivot_for_chart": False,
        "data_type": "Number",
        "multilevel_headers": False,
        "source": {
            "Table ID": "",
            "URL": "https://www.countyhealthrankings.org/",
            "Data Year": 2021,
            "Estimate Type": "",
            "Citation Month": "April",
            "Citation Year": 2025,
            "Custom Text": "County Health Rankings and Roadmaps, 2025",
        },
        "data": [
            {"Year": 2019, "Dutchess": 1500, "Orange": 1450, "Putnam": 2090, "Rockland": 1100,
             "Sullivan": 2900, "Ulster": 1480, "Westchester": 720, "NYS": 1180},
            {"Year": 2020, "Dutchess": 1440, "Orange": 1440, "Putnam": 1970, "Rockland": 1130,
             "Sullivan": 2710, "Ulster": 840, "Westchester": 720, "NYS": 1170},
            {"Year": 2021, "Dutchess": 1410, "Orange": 1430, "Putnam": 1880, "Rockland": 1180,
             "Sullivan": 3070, "Ulster": 1680, "Westchester": 760, "NYS": 1240},
        ],
    },
    # ── Dentists ──────────────────────────────────────────────────────────────
    {
        "sheet_name": "Dentists",
        "object_id": "dentists",
        "caption": "Ratio of Residents to Dentists, 2019-2022",
        "type": "both",
        "figure_type": "Line",
        "x_col": "Year",
        "y_cols": "",
        "y_axis_title": "Ratio of Residents to Dentists",
        "start_at_zero": False,
        "pivot_for_chart": False,
        "data_type": "Number",
        "multilevel_headers": False,
        "source": {
            "Table ID": "",
            "URL": "https://www.countyhealthrankings.org/",
            "Data Year": 2022,
            "Estimate Type": "",
            "Citation Month": "April",
            "Citation Year": 2025,
            "Custom Text": "County Health Rankings and Roadmaps, 2025",
        },
        "data": [
            {"Year": 2019, "Dutchess": 1370, "Orange": 1420, "Putnam": 1670, "Rockland": 980,
             "Sullivan": 2430, "Ulster": 1570, "Westchester": 890, "NYS": 1170},
            {"Year": 2020, "Dutchess": 1380, "Orange": 1460, "Putnam": 1700, "Rockland": 1020,
             "Sullivan": 2370, "Ulster": 1480, "Westchester": 900, "NYS": 1190},
            {"Year": 2021, "Dutchess": 1410, "Orange": 1490, "Putnam": 1660, "Rockland": 1060,
             "Sullivan": 2490, "Ulster": 1490, "Westchester": 930, "NYS": 1220},
            {"Year": 2022, "Dutchess": 1400, "Orange": 1500, "Putnam": 1610, "Rockland": 1060,
             "Sullivan": 2410, "Ulster": 1470, "Westchester": 910, "NYS": 1200},
        ],
    },
    # ── Mental Health Providers ───────────────────────────────────────────────
    {
        "sheet_name": "Mental-Health-Providers",
        "object_id": "mental-health-providers",
        "caption": "Ratio of Residents to Mental Health Providers, 2021-2024",
        "type": "both",
        "figure_type": "Line",
        "x_col": "Year",
        "y_cols": "",
        "y_axis_title": "Ratio of Residents to Mental Health Providers",
        "start_at_zero": False,
        "pivot_for_chart": False,
        "data_type": "Number",
        "multilevel_headers": False,
        "source": {
            "Table ID": "",
            "URL": "https://www.countyhealthrankings.org/",
            "Data Year": 2024,
            "Estimate Type": "",
            "Citation Month": "April",
            "Citation Year": 2025,
            "Custom Text": "County Health Rankings and Roadmaps, 2025",
        },
        "data": [
            {"Year": 2021, "Dutchess": 320, "Orange": 390, "Putnam": 260, "Rockland": 340,
             "Sullivan": 510, "Ulster": 270, "Westchester": 230, "NYS": 310},
            {"Year": 2022, "Dutchess": 310, "Orange": 390, "Putnam": 240, "Rockland": 330,
             "Sullivan": 490, "Ulster": 260, "Westchester": 230, "NYS": 300},
            {"Year": 2023, "Dutchess": 300, "Orange": 370, "Putnam": 230, "Rockland": 300,
             "Sullivan": 450, "Ulster": 250, "Westchester": 220, "NYS": 280},
            {"Year": 2024, "Dutchess": 290, "Orange": 350, "Putnam": 210, "Rockland": 290,
             "Sullivan": 450, "Ulster": 240, "Westchester": 200, "NYS": 260},
        ],
    },
    # ── Access to Primary Care ────────────────────────────────────────────────
    {
        "sheet_name": "Access-Primary-Care",
        "object_id": "access-to-primary-care",
        "caption": "Percentage of Adults Who Have a Regular Primary Care Provider, 2016, 2018, and 2021",
        "type": "both",
        "figure_type": "Clustered Bar",
        "x_col": "Year",
        "y_cols": "",
        "y_axis_title": "Percentage of Adults Who Have a Regular Primary Care Provider (%)",
        "start_at_zero": False,
        "pivot_for_chart": True,
        "data_type": "Percent",
        "multilevel_headers": False,
        "source": {
            "Table ID": "",
            "URL": "https://webbi1.health.ny.gov/SASStoredProcess/guest?_program=%2FEBI%2FPHIG%2Fapps%2Fchir_dashboard%2Fchir_dashboard",
            "Data Year": 2021,
            "Estimate Type": "",
            "Citation Month": "April",
            "Citation Year": 2025,
            "Custom Text": "NYS Community Health Indicator Reports Dashboard, April 2025",
        },
        "data": [
            {"Year": 2016, "Dutchess": 82.8, "Orange": 81.8, "Putnam": 86.7, "Rockland": 84.1,
             "Sullivan": 84.6, "Ulster": 82.5, "Westchester": 79.2, "Mid-Hudson": "s", "NYS": 82.6},
            {"Year": 2018, "Dutchess": 85.7, "Orange": 80.7, "Putnam": 89.0, "Rockland": 83.8,
             "Sullivan": 75.0, "Ulster": 78.3, "Westchester": 81.4, "Mid-Hudson": "s", "NYS": 79.1},
            {"Year": 2021, "Dutchess": 81.6, "Orange": 88.9, "Putnam": 90.5, "Rockland": 84.1,
             "Sullivan": 76.9, "Ulster": 89.5, "Westchester": 84.9, "Mid-Hudson": 86.5, "NYS": 85.0},
        ],
    },
]


def main() -> None:
    from openpyxl import load_workbook

    if not WORKBOOK_PATH.exists():
        raise FileNotFoundError(f"Workbook not found: {WORKBOOK_PATH}")

    print(f"Loading workbook: {WORKBOOK_PATH.name}")
    wb = load_workbook(WORKBOOK_PATH)

    for indicator in PATCH_INDICATORS:
        sheet_name = indicator["sheet_name"]
        # Delete existing sheet if present (needed to recreate with updated config)
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
            print(f"  [del] Deleted existing '{sheet_name}' for recreation.")
        _write_flat_sheet(wb, indicator)
        print(f"  [ok] Created sheet '{sheet_name}'.")

    wb.save(WORKBOOK_PATH)
    print(f"\nSaved: {WORKBOOK_PATH}")
    print(f"Total sheets: {len(wb.sheetnames)}")


if __name__ == "__main__":
    main()
