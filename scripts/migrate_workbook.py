"""
migrate_workbook.py

One-time migration script that brings the main CHA workbook to the fully
automated flat per-sheet format used by workbook_loader.py:

  1. Renames the ``Template`` sheet to ``_Template`` so it is skipped by the
     loader (sheets starting with ``_`` are now ignored).
  2. Removes the vestigial ``data_fig_unemployment`` and
     ``data_tbl_unemployment`` sheets (left over from an older approach).
  3. Fills in Source Specifications for the existing ``Labor-Force`` and
     ``Population Unemployed`` indicator sheets.
  4. Adds all 28 ch04 indicators (sourced from build_content_registry.py data)
     as new flat indicator sheets in the correct format expected by
     workbook_loader._parse_flat_indicator_sheet().

Run once from the project root:
    python scripts/migrate_workbook.py
"""

from __future__ import annotations

import math
from pathlib import Path

import openpyxl
from openpyxl import load_workbook

# ── paths ─────────────────────────────────────────────────────────────────────
PROJECT_ROOT = Path(__file__).parent.parent
WORKBOOK_PATH = PROJECT_ROOT / "data" / "raw" / "Mid-Hudson Regional Community Health Assessment 2025 Data File.xlsx"

# ── flat-format layout constants (1-based openpyxl columns) ──────────────────
COL_KEY = 1    # A – config key
COL_VAL = 2    # B – config value
COL_MARKER = 5 # E – "Enter Data" marker
COL_DATA_START = 6  # F – first data column

# ── section header labels (mirrored from workbook_loader) ────────────────────
SEC_TABLE_RULES  = "Table Rules"
SEC_FIGURE_RULES = "Figure Rules"
SEC_SOURCE_SPECS = "Source Specifications"

# ── source specs for the two existing indicator sheets ───────────────────────
EXISTING_SOURCE_SPECS = {
    "Labor-Force": {
        "Table ID": "DP03",
        "URL": "https://data.census.gov/table/ACSDP5Y2023.DP03",
        "Data Year": 2023,
        "Estimate Type": "5-Year Estimates",
        "Citation Month": "April",
        "Citation Year": 2025,
        "Custom Text": "",
    },
    "Population Unemployed": {
        "Table ID": "DP03",
        "URL": "https://data.census.gov/table/ACSDP5Y2023.DP03",
        "Data Year": 2023,
        "Estimate Type": "5-Year Estimates",
        "Citation Month": "April",
        "Citation Year": 2025,
        "Custom Text": "",
    },
}

# ── ch04 indicator definitions ────────────────────────────────────────────────
# Each entry maps directly to the flat per-sheet format.
# ``type``: "both" | "table"
# ``figure_type``: "Line" | "Clustered Bar" | "Stacked Bar" | "Simple Bar" | "Horizontal Bar" (ignored for table-only)
# ``pivot_for_chart``: True means transpose data before charting (covers both
#     build_content_registry.py's pivot_for_chart=True and transpose_for_chart=True)
CH04_INDICATORS: list[dict] = [
    # 1 ── labor-force (already in workbook; skip if sheet already exists)
    {
        "sheet_name": "Labor-Force",
        "object_id": "labor-force",
        "caption": "Percentage of the Labor Force, Population 16 Years and Older, 2021-2023",
        "type": "both",
        "figure_type": "Line",
        "x_col": "Year",
        "y_cols": "",
        "y_axis_title": "Percentage of Labor Force (%)",
        "start_at_zero": False,
        "pivot_for_chart": True,
        "data_type": "Percent",
        "multilevel_headers": False,
        "source": {"Table ID": "DP03", "URL": "https://data.census.gov/table/ACSDP5Y2023.DP03",
                   "Data Year": 2023, "Estimate Type": "5-Year Estimates",
                   "Citation Month": "April", "Citation Year": 2025, "Custom Text": ""},
        "data": [
            {"Year": 2021, "Dutchess": 63.0, "Orange": 63.9, "Putnam": 65.2, "Rockland": 63.9,
             "Sullivan": 56.1, "Ulster": 60.1, "Westchester": 65.5, "NYS": 63.1, "US": 63.6},
            {"Year": 2022, "Dutchess": 62.8, "Orange": 63.4, "Putnam": 64.7, "Rockland": 63.2,
             "Sullivan": 58.1, "Ulster": 58.9, "Westchester": 65.2, "NYS": 62.9, "US": 63.5},
            {"Year": 2023, "Dutchess": 63.3, "Orange": 63.5, "Putnam": 64.7, "Rockland": 63.2,
             "Sullivan": 59.1, "Ulster": 58.7, "Westchester": 65.4, "NYS": 63.0, "US": 63.5},
        ],
    },
    # 2 ── unemployment (already in workbook as "Population Unemployed")
    {
        "sheet_name": "Population Unemployed",
        "object_id": "unemployed",
        "caption": "Percentage of Population Unemployed, 16 Years and Older, 2021-2023",
        "type": "both",
        "figure_type": "Clustered Bar",
        "x_col": "Year",
        "y_cols": "",
        "y_axis_title": "Percentage Unemployed (%)",
        "start_at_zero": False,
        "pivot_for_chart": True,
        "data_type": "Percent",
        "multilevel_headers": False,
        "source": {"Table ID": "DP03", "URL": "https://data.census.gov/table/ACSDP5Y2023.DP03",
                   "Data Year": 2023, "Estimate Type": "5-Year Estimates",
                   "Citation Month": "April", "Citation Year": 2025, "Custom Text": ""},
        "data": [
            {"Year": 2021, "Dutchess": 5.2, "Orange": 5.4, "Putnam": 4.8, "Rockland": 6.0,
             "Sullivan": 8.0, "Ulster": 5.1, "Westchester": 6.1, "NYS": 6.2, "US": 5.5},
            {"Year": 2022, "Dutchess": 5.0, "Orange": 5.3, "Putnam": 4.4, "Rockland": 6.2,
             "Sullivan": 7.2, "Ulster": 5.1, "Westchester": 6.0, "NYS": 6.2, "US": 5.3},
            {"Year": 2023, "Dutchess": 4.8, "Orange": 5.4, "Putnam": 4.1, "Rockland": 5.9,
             "Sullivan": 6.1, "Ulster": 5.1, "Westchester": 6.0, "NYS": 6.2, "US": 5.2},
        ],
    },
    # 3 ── food-insecurity
    {
        "sheet_name": "Food-Insecurity",
        "object_id": "food-insecurity",
        "caption": "Percentage of Overall Food Insecurity, 2020-2023",
        "type": "both",
        "figure_type": "Line",
        "x_col": "Year",
        "y_cols": "",
        "y_axis_title": "Food Insecurity Rate (%)",
        "start_at_zero": False,
        "pivot_for_chart": False,
        "data_type": "Percent",
        "multilevel_headers": False,
        "source": {"Table ID": "", "URL": "https://map.feedingamerica.org/district/2023/overall/new-york",
                   "Data Year": 2023, "Estimate Type": "",
                   "Citation Month": "June", "Citation Year": 2025, "Custom Text": "Feeding America, June 2025"},
        "data": [
            {"Year": 2020, "Dutchess": 8.7, "Orange": 9.4, "Putnam": 6.3, "Rockland": 9.7,
             "Sullivan": 11.5, "Ulster": 11.3, "Westchester": 7.9, "NYS": 9.6},
            {"Year": 2021, "Dutchess": 7.2, "Orange": 7.8, "Putnam": 5.3, "Rockland": 8.2,
             "Sullivan": 9.9, "Ulster": 9.5, "Westchester": 6.6, "NYS": 11.4},
            {"Year": 2022, "Dutchess": 10.0, "Orange": 11.3, "Putnam": 8.4, "Rockland": 11.0,
             "Sullivan": 13.1, "Ulster": 12.8, "Westchester": 9.4, "NYS": 13.4},
            {"Year": 2023, "Dutchess": 10.8, "Orange": 12.1, "Putnam": 8.9, "Rockland": 12.0,
             "Sullivan": 14.0, "Ulster": 13.2, "Westchester": 10.7, "NYS": 14.5},
        ],
    },
    # 4 ── child-food-insecurity
    {
        "sheet_name": "Child-Food-Insecurity",
        "object_id": "childhood-food-insecurity",
        "caption": "Percentage of Food Insecurity, Children 18 Years and Younger, 2020-2023",
        "type": "both",
        "figure_type": "Line",
        "x_col": "Year",
        "y_cols": "",
        "y_axis_title": "Child Food Insecurity Rate (%)",
        "start_at_zero": False,
        "pivot_for_chart": False,
        "data_type": "Percent",
        "multilevel_headers": False,
        "source": {"Table ID": "", "URL": "https://map.feedingamerica.org/district/2023/overall/new-york",
                   "Data Year": 2023, "Estimate Type": "",
                   "Citation Month": "June", "Citation Year": 2025, "Custom Text": "Feeding America, June 2025"},
        "data": [
            {"Year": 2020, "Dutchess": 11.5, "Orange": 13.0, "Putnam": 7.8, "Rockland": 13.5,
             "Sullivan": 16.5, "Ulster": 15.8, "Westchester": 10.2, "NYS": 14.0},
            {"Year": 2021, "Dutchess": 9.8, "Orange": 11.0, "Putnam": 6.5, "Rockland": 11.5,
             "Sullivan": 14.2, "Ulster": 13.5, "Westchester": 8.9, "NYS": 16.0},
            {"Year": 2022, "Dutchess": 13.5, "Orange": 15.8, "Putnam": 9.8, "Rockland": 15.5,
             "Sullivan": 18.5, "Ulster": 17.8, "Westchester": 12.8, "NYS": 19.0},
            {"Year": 2023, "Dutchess": 14.8, "Orange": 16.8, "Putnam": 6.6, "Rockland": 16.5,
             "Sullivan": 19.9, "Ulster": 18.5, "Westchester": 14.0, "NYS": 20.5},
        ],
    },
    # 5 ── cost-burdened-renters
    {
        "sheet_name": "Cost-Burdened-Renters",
        "object_id": "cost-burdened-renters",
        "caption": "Percentage of Cost Burdened Renter Occupied Units, 2021-2023",
        "type": "both",
        "figure_type": "Clustered Bar",
        "x_col": "Year",
        "y_cols": "",
        "y_axis_title": "Percentage of Cost Burdened Renter Occupied Units (%)",
        "start_at_zero": False,
        "pivot_for_chart": True,
        "data_type": "Percent",
        "multilevel_headers": False,
        "source": {"Table ID": "DP04", "URL": "https://data.census.gov/table/ACSDP5Y2023.DP04",
                   "Data Year": 2023, "Estimate Type": "5-Year Estimates",
                   "Citation Month": "April", "Citation Year": 2025, "Custom Text": ""},
        "data": [
            {"Year": 2021, "Dutchess": 52.3, "Orange": 56.5, "Putnam": 52.7, "Rockland": 58.3,
             "Sullivan": 48.0, "Ulster": 55.3, "Westchester": 53.2, "NYS": 51.6},
            {"Year": 2022, "Dutchess": 52.4, "Orange": 56.1, "Putnam": 53.2, "Rockland": 58.9,
             "Sullivan": 48.4, "Ulster": 55.3, "Westchester": 53.5, "NYS": 51.7},
            {"Year": 2023, "Dutchess": 52.0, "Orange": 56.2, "Putnam": 56.5, "Rockland": 59.6,
             "Sullivan": 48.7, "Ulster": 56.6, "Westchester": 53.0, "NYS": 51.5},
        ],
    },
    # 6 ── severely-cost-burdened
    {
        "sheet_name": "Severely-Cost-Burdened",
        "object_id": "severely-cost-burdened",
        "caption": "Percentage of Severely Cost Burdened Households, 2016-2023",
        "type": "both",
        "figure_type": "Clustered Bar",
        "x_col": "Period",
        "y_cols": "",
        "y_axis_title": "Percentage of Severely Cost Burdened Households (%)",
        "start_at_zero": False,
        "pivot_for_chart": True,
        "data_type": "Percent",
        "multilevel_headers": False,
        "source": {"Table ID": "", "URL": "https://www.countyhealthrankings.org/",
                   "Data Year": 2023, "Estimate Type": "",
                   "Citation Month": "April", "Citation Year": 2025,
                   "Custom Text": "County Health Rankings and Roadmaps, 2025"},
        "data": [
            {"Period": "2016\u20132020", "Dutchess": 16.0, "Orange": 19.0, "Putnam": 16.0,
             "Rockland": 22.0, "Sullivan": 14.0, "Ulster": 18.0, "Westchester": 20.0, "NYS": 19.0},
            {"Period": "2017\u20132021", "Dutchess": 16.0, "Orange": 19.0, "Putnam": 16.0,
             "Rockland": 22.0, "Sullivan": 15.0, "Ulster": 18.0, "Westchester": 19.0, "NYS": 19.0},
            {"Period": "2018\u20132022", "Dutchess": 16.0, "Orange": 20.0, "Putnam": 16.0,
             "Rockland": 22.0, "Sullivan": 15.0, "Ulster": 18.0, "Westchester": 20.0, "NYS": 19.0},
            {"Period": "2019\u20132023", "Dutchess": 16.0, "Orange": 20.0, "Putnam": 17.0,
             "Rockland": 22.0, "Sullivan": 15.0, "Ulster": 18.0, "Westchester": 19.0, "NYS": 19.0},
        ],
    },
    # 7 ── hud-housing
    {
        "sheet_name": "HUD-Housing",
        "object_id": "hud-housing",
        "caption": "Number of People Living in HUD-Subsidized Housing in the Past 12 Months, 2021\u20132024",
        "type": "both",
        "figure_type": "Clustered Bar",
        "x_col": "Year",
        "y_cols": "",
        "y_axis_title": "Number of People",
        "start_at_zero": True,
        "pivot_for_chart": True,
        "data_type": "Number",
        "multilevel_headers": False,
        "source": {"Table ID": "", "URL": "https://www.huduser.gov/portal/datasets/assthsg.html",
                   "Data Year": 2024, "Estimate Type": "",
                   "Citation Month": "July", "Citation Year": 2025,
                   "Custom Text": "US Department of Housing and Urban Development, July 2025"},
        "data": [
            {"Year": 2021, "Dutchess": 7442, "Orange": 18258, "Putnam": 945, "Rockland": 21732,
             "Sullivan": 5018, "Ulster": 5484, "Westchester": 40230, "NYS": 1025652},
            {"Year": 2022, "Dutchess": 7641, "Orange": 18745, "Putnam": 929, "Rockland": 22170,
             "Sullivan": 5228, "Ulster": 5479, "Westchester": 40412, "NYS": 985104},
            {"Year": 2023, "Dutchess": 7630, "Orange": 19000, "Putnam": 992, "Rockland": 23411,
             "Sullivan": 4846, "Ulster": 5418, "Westchester": 40415, "NYS": 987957},
            {"Year": 2024, "Dutchess": 7484, "Orange": 19129, "Putnam": 1012, "Rockland": 23735,
             "Sullivan": 4445, "Ulster": 4980, "Westchester": 40137, "NYS": 1000730},
        ],
    },
    # 8 ── poverty-threshold (table only, multilevel headers)
    {
        "sheet_name": "Poverty-Threshold",
        "object_id": "poverty-threshold",
        "caption": "Poverty Threshold for 2024 by Size of Family and Number of Related Children 18 Years and Younger",
        "type": "table",
        "figure_type": "",
        "x_col": "Family Size",
        "y_cols": "",
        "y_axis_title": "",
        "start_at_zero": False,
        "pivot_for_chart": False,
        "data_type": "Number",
        "multilevel_headers": True,
        "source": {"Table ID": "", "URL": "https://www.census.gov/data/tables/time-series/demo/income-poverty/historical-poverty-thresholds.html",
                   "Data Year": 2024, "Estimate Type": "",
                   "Citation Month": "April", "Citation Year": 2025,
                   "Custom Text": "US Census Bureau, 2024 Poverty Thresholds, April 2025"},
        "data": [
            {"Family Size": "One person (unrelated individual): Under age 65",
             "None": "$16,320", "One": "", "Two": "", "Three": "", "Four": "",
             "Five": "", "Six": "", "Seven": "", "Eight": "", "Nine or more": ""},
            {"Family Size": "One person (unrelated individual): Aged 65 and older",
             "None": "$15,045", "One": "", "Two": "", "Three": "", "Four": "",
             "Five": "", "Six": "", "Seven": "", "Eight": "", "Nine or more": ""},
            {"Family Size": "Two people: Householder under age 65",
             "None": "$21,006", "One": "$21,621", "Two": "", "Three": "", "Four": "",
             "Five": "", "Six": "", "Seven": "", "Eight": "", "Nine or more": ""},
            {"Family Size": "Two people: Householder aged 65 and older",
             "None": "$18,961", "One": "$21,540", "Two": "", "Three": "", "Four": "",
             "Five": "", "Six": "", "Seven": "", "Eight": "", "Nine or more": ""},
            {"Family Size": "Three people",
             "None": "$24,537", "One": "$25,249", "Two": "$25,273", "Three": "", "Four": "",
             "Five": "", "Six": "", "Seven": "", "Eight": "", "Nine or more": ""},
            {"Family Size": "Four people",
             "None": "$32,355", "One": "$32,884", "Two": "$31,812", "Three": "$31,922",
             "Four": "", "Five": "", "Six": "", "Seven": "", "Eight": "", "Nine or more": ""},
            {"Family Size": "Five people",
             "None": "$39,019", "One": "$39,586", "Two": "$38,374", "Three": "$37,436",
             "Four": "$36,863", "Five": "", "Six": "", "Seven": "", "Eight": "", "Nine or more": ""},
            {"Family Size": "Six people",
             "None": "$44,879", "One": "$45,057", "Two": "$44,128", "Three": "$43,238",
             "Four": "$41,915", "Five": "$41,131", "Six": "", "Seven": "", "Eight": "", "Nine or more": ""},
            {"Family Size": "Seven people",
             "None": "$51,638", "One": "$51,961", "Two": "$50,849", "Three": "$50,075",
             "Four": "$48,631", "Five": "$46,948", "Six": "$45,100", "Seven": "", "Eight": "", "Nine or more": ""},
            {"Family Size": "Eight people",
             "None": "$57,753", "One": "$58,263", "Two": "$57,215", "Three": "$56,296",
             "Four": "$54,992", "Five": "$53,337", "Six": "$51,614", "Seven": "$51,177",
             "Eight": "", "Nine or more": ""},
            {"Family Size": "Nine people or more",
             "None": "$69,473", "One": "$69,810", "Two": "$68,882", "Three": "$68,102",
             "Four": "$66,822", "Five": "$65,062", "Six": "$63,469", "Seven": "$63,075",
             "Eight": "$60,645", "Nine or more": "$60,645"},
        ],
    },
    # 9 ── poverty-rate
    {
        "sheet_name": "Poverty-Rate",
        "object_id": "poverty-rate",
        "caption": "Percentage of Population in Poverty, 2021-2023",
        "type": "both",
        "figure_type": "Clustered Bar",
        "x_col": "Year",
        "y_cols": "",
        "y_axis_title": "Poverty Rate (%)",
        "start_at_zero": True,
        "pivot_for_chart": True,
        "data_type": "Percent",
        "multilevel_headers": False,
        "source": {"Table ID": "S1701", "URL": "https://data.census.gov/table/ACSST5Y2023.S1701",
                   "Data Year": 2023, "Estimate Type": "5-Year Estimates",
                   "Citation Month": "April", "Citation Year": 2025, "Custom Text": ""},
        "data": [
            {"Year": 2021, "Dutchess": 8.8, "Orange": 11.7, "Putnam": 6.0, "Rockland": 14.9,
             "Sullivan": 14.1, "Ulster": 13.2, "Westchester": 8.2, "NYS": 13.5},
            {"Year": 2022, "Dutchess": 8.6, "Orange": 13.0, "Putnam": 6.3, "Rockland": 15.1,
             "Sullivan": 14.8, "Ulster": 14.7, "Westchester": 8.5, "NYS": 13.6},
            {"Year": 2023, "Dutchess": 8.3, "Orange": 13.0, "Putnam": 6.5, "Rockland": 15.6,
             "Sullivan": 15.2, "Ulster": 14.3, "Westchester": 8.9, "NYS": 13.7},
        ],
    },
    # 10 ── poverty-by-race
    {
        "sheet_name": "Poverty-By-Race",
        "object_id": "poverty-by-race",
        "caption": "Percentage of Families Below Poverty Level by Race and Ethnicity, 2023",
        "type": "both",
        "figure_type": "Clustered Bar",
        "x_col": "Race/Ethnicity",
        "y_cols": "White (non-Hispanic),Black (including Hispanic),Asian (including Hispanic, excluding PI),Hispanic (any race),Total",
        "y_axis_title": "Poverty Rate (%)",
        "start_at_zero": True,
        "pivot_for_chart": True,
        "data_type": "Percent",
        "multilevel_headers": False,
        "source": {"Table ID": "S1702", "URL": "https://data.census.gov/table/ACSST5Y2023.S1702",
                   "Data Year": 2023, "Estimate Type": "5-Year Estimates",
                   "Citation Month": "April", "Citation Year": 2025, "Custom Text": ""},
        "data": [
            {"Race/Ethnicity": "White (non-Hispanic)", "Dutchess": 3.3, "Orange": 7.9,
             "Putnam": 2.8, "Rockland": 11.1, "Sullivan": 8.2, "Ulster": 4.9, "Westchester": 2.6, "NYS": 5.7},
            {"Race/Ethnicity": "Black (including Hispanic)", "Dutchess": 14.2, "Orange": 12.5,
             "Putnam": 0.0, "Rockland": 8.6, "Sullivan": 30.5, "Ulster": 14.1, "Westchester": 10.8, "NYS": 16.5},
            {"Race/Ethnicity": "Asian (including Hispanic, excluding PI)", "Dutchess": 7.5,
             "Orange": 10.3, "Putnam": 23.4, "Rockland": 1.8, "Sullivan": 13.3, "Ulster": 4.8,
             "Westchester": 5.4, "NYS": 11.1},
            {"Race/Ethnicity": "Hispanic (any race)", "Dutchess": 11.2, "Orange": 11.3,
             "Putnam": 6.6, "Rockland": 12.8, "Sullivan": 14.3, "Ulster": 24.4,
             "Westchester": 10.4, "NYS": 17.1},
            {"Race/Ethnicity": "Total", "Dutchess": 5.3, "Orange": 9.2, "Putnam": 3.8,
             "Rockland": 10.3, "Sullivan": 10.5, "Ulster": 7.1, "Westchester": 5.7, "NYS": 9.7},
        ],
    },
    # 11 ── economically-disadvantaged
    {
        "sheet_name": "Economically-Disadvantaged",
        "object_id": "economically-disadvantaged",
        "caption": "Enrollment Rate of Economically Disadvantaged Students, 2021-2024",
        "type": "both",
        "figure_type": "Line",
        "x_col": "School Year",
        "y_cols": "",
        "y_axis_title": "Economically Disadvantaged Students (%)",
        "start_at_zero": False,
        "pivot_for_chart": False,
        "data_type": "Percent",
        "multilevel_headers": False,
        "source": {"Table ID": "", "URL": "https://data.nysed.gov/",
                   "Data Year": 2024, "Estimate Type": "",
                   "Citation Month": "April", "Citation Year": 2025,
                   "Custom Text": "NYS Education Department, April 2025"},
        "data": [
            {"School Year": "2021-2022", "Dutchess": 39.2, "Orange": 47.0, "Putnam": 28.5,
             "Rockland": 43.9, "Sullivan": 60.4, "Ulster": 48.6, "Westchester": 38.9, "NYS": 56.2},
            {"School Year": "2022-2023", "Dutchess": 43.1, "Orange": 49.0, "Putnam": 32.1,
             "Rockland": 48.1, "Sullivan": 62.7, "Ulster": 52.5, "Westchester": 39.6, "NYS": 59.1},
            {"School Year": "2023-2024", "Dutchess": 43.3, "Orange": 49.5, "Putnam": 32.5,
             "Rockland": 51.0, "Sullivan": 61.4, "Ulster": 48.8, "Westchester": 39.7, "NYS": 59.2},
        ],
    },
    # 12 ── alice-budget (table only)
    {
        "sheet_name": "ALICE-Budget",
        "object_id": "alice-budget",
        "caption": "ALICE Household Survival Budget, New York State, 2023",
        "type": "table",
        "figure_type": "",
        "x_col": "Budget Item",
        "y_cols": "",
        "y_axis_title": "",
        "start_at_zero": False,
        "pivot_for_chart": False,
        "data_type": "Number",
        "multilevel_headers": False,
        "source": {"Table ID": "", "URL": "https://www.unitedforalice.org/county-reports/new-york",
                   "Data Year": 2023, "Estimate Type": "",
                   "Citation Month": "April", "Citation Year": 2025,
                   "Custom Text": "United For ALICE, 2023 County Reports, April 2025"},
        "data": [
            {"Budget Item": "Housing", "Single Adult": 1103, "One Adult, One Child": 1189,
             "One Adult, One in Child Care": 1189, "Two Adults": 1189,
             "Two Adults, Two Children": 1437, "Two Adults, Two in Child Care": 1437,
             "Single Adult 65+": 1103, "Two Adults 65+": 1189},
            {"Budget Item": "Child Care", "Single Adult": 0, "One Adult, One Child": 423,
             "One Adult, One in Child Care": 1126, "Two Adults": 0,
             "Two Adults, Two Children": 844, "Two Adults, Two in Child Care": 2345,
             "Single Adult 65+": 0, "Two Adults 65+": 0},
            {"Budget Item": "Food", "Single Adult": 516, "One Adult, One Child": 873,
             "One Adult, One in Child Care": 783, "Two Adults": 946,
             "Two Adults, Two Children": 1587, "Two Adults, Two in Child Care": 1400,
             "Single Adult 65+": 475, "Two Adults 65+": 870},
            {"Budget Item": "Transportation", "Single Adult": 401, "One Adult, One Child": 535,
             "One Adult, One in Child Care": 506, "Two Adults": 617,
             "Two Adults, Two Children": 957, "Two Adults, Two in Child Care": 899,
             "Single Adult 65+": 346, "Two Adults 65+": 508},
            {"Budget Item": "Health Care", "Single Adult": 196, "One Adult, One Child": 452,
             "One Adult, One in Child Care": 452, "Two Adults": 452,
             "Two Adults, Two Children": 775, "Two Adults, Two in Child Care": 775,
             "Single Adult 65+": 543, "Two Adults 65+": 1086},
            {"Budget Item": "Technology", "Single Adult": 86, "One Adult, One Child": 86,
             "One Adult, One in Child Care": 86, "Two Adults": 116,
             "Two Adults, Two Children": 116, "Two Adults, Two in Child Care": 116,
             "Single Adult 65+": 86, "Two Adults 65+": 116},
            {"Budget Item": "Miscellaneous", "Single Adult": 230, "One Adult, One Child": 356,
             "One Adult, One in Child Care": 414, "Two Adults": 332,
             "Two Adults, Two Children": 572, "Two Adults, Two in Child Care": 697,
             "Single Adult 65+": 255, "Two Adults 65+": 377},
            {"Budget Item": "Taxes", "Single Adult": 439, "One Adult, One Child": 461,
             "One Adult, One in Child Care": 625, "Two Adults": 529,
             "Two Adults, Two Children": 684, "Two Adults, Two in Child Care": 1037,
             "Single Adult 65+": 510, "Two Adults 65+": 861},
            {"Budget Item": "Monthly Total", "Single Adult": 2971, "One Adult, One Child": 4375,
             "One Adult, One in Child Care": 5181, "Two Adults": 4181,
             "Two Adults, Two Children": 6972, "Two Adults, Two in Child Care": 8706,
             "Single Adult 65+": 3318, "Two Adults 65+": 5007},
            {"Budget Item": "ANNUAL TOTAL", "Single Adult": 35652, "One Adult, One Child": 52500,
             "One Adult, One in Child Care": 62172, "Two Adults": 50172,
             "Two Adults, Two Children": 83664, "Two Adults, Two in Child Care": 104472,
             "Single Adult 65+": 39816, "Two Adults 65+": 60084},
            {"Budget Item": "Hourly Wage", "Single Adult": 17.83, "One Adult, One Child": 26.25,
             "One Adult, One in Child Care": 31.09, "Two Adults": 25.09,
             "Two Adults, Two Children": 41.83, "Two Adults, Two in Child Care": 52.24,
             "Single Adult 65+": 19.91, "Two Adults 65+": 30.04},
        ],
    },
    # 13 ── alice-threshold
    {
        "sheet_name": "ALICE-Threshold",
        "object_id": "alice-threshold",
        "caption": "ALICE Threshold Percentage, 2023",
        "type": "both",
        "figure_type": "Stacked Bar",
        "x_col": "Category",
        "y_cols": "ALICE,Poverty,Above ALICE Threshold",
        "y_axis_title": "Percentage (%)",
        "start_at_zero": True,
        "pivot_for_chart": True,
        "data_type": "Percent",
        "multilevel_headers": False,
        "source": {"Table ID": "", "URL": "https://www.unitedforalice.org/county-reports/new-york",
                   "Data Year": 2023, "Estimate Type": "",
                   "Citation Month": "April", "Citation Year": 2025,
                   "Custom Text": "United For ALICE, 2023 County Reports, April 2025"},
        "data": [
            {"Category": "ALICE", "Dutchess": 31.0, "Orange": 33.0, "Putnam": 32.0,
             "Rockland": 38.0, "Sullivan": 30.0, "Ulster": 32.0, "Westchester": 28.0, "NYS": 33.0},
            {"Category": "Poverty", "Dutchess": 6.0, "Orange": 10.0, "Putnam": 6.0,
             "Rockland": 12.0, "Sullivan": 15.0, "Ulster": 10.0, "Westchester": 10.0, "NYS": 14.0},
            {"Category": "Above ALICE Threshold", "Dutchess": 62.0, "Orange": 57.0, "Putnam": 63.0,
             "Rockland": 50.0, "Sullivan": 54.0, "Ulster": 58.0, "Westchester": 61.0, "NYS": 52.0},
        ],
    },
    # 14 ── graduation-rate
    {
        "sheet_name": "Graduation-Rate",
        "object_id": "graduation-rate",
        "caption": "High School Graduation Rate, 2021-2023",
        "type": "both",
        "figure_type": "Clustered Bar",
        "x_col": "Year",
        "y_cols": "",
        "y_axis_title": "High School Graduation Rate (%)",
        "start_at_zero": False,
        "pivot_for_chart": True,
        "data_type": "Percent",
        "multilevel_headers": False,
        "source": {"Table ID": "", "URL": "https://data.nysed.gov/gradrate.php",
                   "Data Year": 2023, "Estimate Type": "",
                   "Citation Month": "April", "Citation Year": 2025,
                   "Custom Text": "NYS Education Department, Graduation Rate Data, April 2025"},
        "data": [
            {"Year": 2021, "Dutchess": 87.0, "Orange": 89.0, "Putnam": 94.0, "Rockland": 90.0,
             "Sullivan": 82.0, "Ulster": 87.0, "Westchester": 91.0, "NYS": 86.0},
            {"Year": 2022, "Dutchess": 86.0, "Orange": 89.0, "Putnam": 94.0, "Rockland": 88.0,
             "Sullivan": 78.0, "Ulster": 87.0, "Westchester": 92.0, "NYS": 87.0},
            {"Year": 2023, "Dutchess": 87.0, "Orange": 89.0, "Putnam": 91.0, "Rockland": 86.0,
             "Sullivan": 76.0, "Ulster": 87.0, "Westchester": 91.0, "NYS": 86.0},
        ],
    },
    # 15 ── graduation-by-race
    {
        "sheet_name": "Graduation-By-Race",
        "object_id": "graduation-by-race",
        "caption": "High School Graduation Rate, by Race and Ethnicity, 2023",
        "type": "both",
        "figure_type": "Clustered Bar",
        "x_col": "Race/Ethnicity",
        "y_cols": "",
        "y_axis_title": "High School Graduation Rate (%)",
        "start_at_zero": True,
        "pivot_for_chart": False,
        "data_type": "Percent",
        "multilevel_headers": False,
        "source": {"Table ID": "", "URL": "https://data.nysed.gov/gradrate.php",
                   "Data Year": 2023, "Estimate Type": "",
                   "Citation Month": "April", "Citation Year": 2025,
                   "Custom Text": "NYS Education Department, Graduation Rate Data, April 2025"},
        "data": [
            {"Race/Ethnicity": "Asian or Native Hawaiian/Other PI",
             "Dutchess": 91.0, "Orange": 96.0, "Putnam": 100.0, "Rockland": 97.0,
             "Sullivan": "s", "Ulster": "s", "Westchester": 97.0, "NYS": 93.0},
            {"Race/Ethnicity": "Black or African American",
             "Dutchess": 78.0, "Orange": 85.0, "Putnam": 97.0, "Rockland": 86.0,
             "Sullivan": 66.0, "Ulster": 84.0, "Westchester": 83.0, "NYS": 81.0},
            {"Race/Ethnicity": "White",
             "Dutchess": 92.0, "Orange": 93.0, "Putnam": 95.0, "Rockland": 94.0,
             "Sullivan": 84.0, "Ulster": 90.0, "Westchester": 96.0, "NYS": 91.0},
            {"Race/Ethnicity": "Multiracial",
             "Dutchess": 80.0, "Orange": 88.0, "Putnam": "s", "Rockland": "s",
             "Sullivan": 68.0, "Ulster": 74.0, "Westchester": 94.0, "NYS": 84.0},
            {"Race/Ethnicity": "Hispanic",
             "Dutchess": 79.0, "Orange": 83.0, "Putnam": 80.0, "Rockland": 74.0,
             "Sullivan": 63.0, "Ulster": 80.0, "Westchester": 85.0, "NYS": 81.0},
        ],
    },
    # 16 ── educational-attainment
    {
        "sheet_name": "Educational-Attainment",
        "object_id": "educational-attainment",
        "caption": "Rate of Education Attainment, 2023",
        "type": "both",
        "figure_type": "Stacked Bar",
        "x_col": "Education Level",
        "y_cols": "",
        "y_axis_title": "Percentage (%)",
        "start_at_zero": True,
        "pivot_for_chart": False,
        "data_type": "Percent",
        "multilevel_headers": False,
        "source": {"Table ID": "S1501", "URL": "https://data.census.gov/table/ACSST5Y2023.S1501",
                   "Data Year": 2023, "Estimate Type": "5-Year Estimates",
                   "Citation Month": "April", "Citation Year": 2025, "Custom Text": ""},
        "data": [
            {"Education Level": "Some college, no degree",
             "Dutchess": 16.9, "Orange": 19.1, "Putnam": 16.5, "Rockland": 16.2,
             "Sullivan": 17.9, "Ulster": 17.1, "Westchester": 12.6, "NYS": 14.9},
            {"Education Level": "Associate's degree",
             "Dutchess": 10.1, "Orange": 10.3, "Putnam": 8.4, "Rockland": 7.8,
             "Sullivan": 10.3, "Ulster": 9.6, "Westchester": 6.4, "NYS": 8.9},
            {"Education Level": "Bachelor's degree",
             "Dutchess": 21.0, "Orange": 18.2, "Putnam": 24.2, "Rockland": 23.2,
             "Sullivan": 17.0, "Ulster": 20.2, "Westchester": 26.0, "NYS": 22.0},
            {"Education Level": "Graduate or professional degree",
             "Dutchess": 18.7, "Orange": 14.1, "Putnam": 19.3, "Rockland": 18.9,
             "Sullivan": 12.7, "Ulster": 16.9, "Westchester": 26.5, "NYS": 17.5},
        ],
    },
    # 17 ── language-proficiency
    {
        "sheet_name": "Language-Proficiency",
        "object_id": "language-proficiency",
        "caption": "Percentage of Population that Speaks English Less Than Very Well, 5 Years and Older, 2021-2023",
        "type": "both",
        "figure_type": "Clustered Bar",
        "x_col": "Year",
        "y_cols": "",
        "y_axis_title": "Percentage Speaking English Less Than Very Well (%)",
        "start_at_zero": False,
        "pivot_for_chart": True,
        "data_type": "Percent",
        "multilevel_headers": False,
        "source": {"Table ID": "S1601", "URL": "https://data.census.gov/table/ACSST5Y2023.S1601",
                   "Data Year": 2023, "Estimate Type": "5-Year Estimates",
                   "Citation Month": "April", "Citation Year": 2025, "Custom Text": ""},
        "data": [
            {"Year": 2021, "Dutchess": 4.7, "Orange": 10.6, "Putnam": 5.3, "Rockland": 18.6,
             "Sullivan": 5.9, "Ulster": 2.9, "Westchester": 12.7, "NYS": 13.1},
            {"Year": 2022, "Dutchess": 4.8, "Orange": 11.3, "Putnam": 5.6, "Rockland": 18.8,
             "Sullivan": 7.3, "Ulster": 3.0, "Westchester": 12.2, "NYS": 13.1},
            {"Year": 2023, "Dutchess": 5.4, "Orange": 12.2, "Putnam": 6.0, "Rockland": 20.0,
             "Sullivan": 8.0, "Ulster": 3.4, "Westchester": 12.4, "NYS": 13.3},
        ],
    },
    # 18 ── disconnected-youth
    {
        "sheet_name": "Disconnected-Youth",
        "object_id": "disconnected-youth",
        "caption": "Percentage of Disconnected Youth Ages 16-19, 2014-2023",
        "type": "both",
        "figure_type": "Clustered Bar",
        "x_col": "Period",
        "y_cols": "",
        "y_axis_title": "Percentage of Disconnected Youth (%)",
        "start_at_zero": False,
        "pivot_for_chart": True,
        "data_type": "Percent",
        "multilevel_headers": False,
        "source": {"Table ID": "S1401", "URL": "https://data.census.gov/table/ACSST5Y2023.S1401",
                   "Data Year": 2023, "Estimate Type": "5-Year Estimates",
                   "Citation Month": "April", "Citation Year": 2025, "Custom Text": ""},
        "data": [
            {"Period": "2014\u20132018", "Dutchess": 4.0, "Orange": 8.0, "Putnam": 4.0,
             "Rockland": 5.0, "Sullivan": 12.0, "Ulster": 6.0, "Westchester": 6.0, "NYS": 6.0},
            {"Period": "2015\u20132019", "Dutchess": 5.0, "Orange": 8.0, "Putnam": "s",
             "Rockland": 5.0, "Sullivan": 12.0, "Ulster": 6.0, "Westchester": 6.0, "NYS": 6.0},
            {"Period": "2016\u20132020", "Dutchess": 5.0, "Orange": 10.0, "Putnam": "s",
             "Rockland": 5.0, "Sullivan": 17.0, "Ulster": 6.0, "Westchester": 4.0, "NYS": 6.0},
            {"Period": "2017\u20132021", "Dutchess": 6.0, "Orange": 8.0, "Putnam": "s",
             "Rockland": 4.0, "Sullivan": 17.0, "Ulster": 6.0, "Westchester": 4.0, "NYS": 6.0},
            {"Period": "2018\u20132022", "Dutchess": 7.0, "Orange": 8.0, "Putnam": "s",
             "Rockland": 6.0, "Sullivan": 25.0, "Ulster": 9.0, "Westchester": 5.0, "NYS": 7.0},
            {"Period": "2019\u20132023", "Dutchess": 6.0, "Orange": 8.0, "Putnam": "s",
             "Rockland": 6.0, "Sullivan": 24.0, "Ulster": 9.0, "Westchester": 5.0, "NYS": 7.0},
        ],
    },
    # 19 ── residential-segregation (table only)
    {
        "sheet_name": "Residential-Segregation",
        "object_id": "residential-segregation",
        "caption": "Index Score of Residential Segregation, 2013-2023",
        "type": "table",
        "figure_type": "",
        "x_col": "Period",
        "y_cols": "",
        "y_axis_title": "",
        "start_at_zero": False,
        "pivot_for_chart": False,
        "data_type": "Number",
        "multilevel_headers": False,
        "source": {"Table ID": "", "URL": "https://www.countyhealthrankings.org/",
                   "Data Year": 2023, "Estimate Type": "",
                   "Citation Month": "April", "Citation Year": 2025,
                   "Custom Text": "County Health Rankings and Roadmaps, 2025"},
        "data": [
            {"Period": "2013\u20132017", "Dutchess": 52.0, "Orange": 44.0, "Putnam": 39.0,
             "Rockland": 58.0, "Sullivan": 46.0, "Ulster": 49.0, "Westchester": 62.0, "NYS": 74.0},
            {"Period": "2016\u20132020", "Dutchess": 50.0, "Orange": 45.0, "Putnam": 44.0,
             "Rockland": 55.0, "Sullivan": 50.0, "Ulster": 50.0, "Westchester": 59.0, "NYS": 74.0},
            {"Period": "2017\u20132021", "Dutchess": 47.0, "Orange": 47.0, "Putnam": 38.0,
             "Rockland": 58.0, "Sullivan": 55.0, "Ulster": 46.0, "Westchester": 59.0, "NYS": 74.0},
            {"Period": "2018\u20132022", "Dutchess": 46.0, "Orange": 49.0, "Putnam": 45.0,
             "Rockland": 56.0, "Sullivan": 51.0, "Ulster": 43.0, "Westchester": 60.0, "NYS": 74.0},
            {"Period": "2019\u20132023", "Dutchess": 49.0, "Orange": 48.0, "Putnam": 49.0,
             "Rockland": 58.0, "Sullivan": 43.0, "Ulster": "s",
             "Westchester": 63.0, "NYS": 75.0},
        ],
    },
    # 20 ── food-environment-index
    {
        "sheet_name": "Food-Environment-Index",
        "object_id": "food-environment-index",
        "caption": "Food Environment Index, 2025",
        "type": "both",
        "figure_type": "Clustered Bar",
        "x_col": "County",
        "y_cols": "Food Environment Index",
        "y_axis_title": "Food Environment Index (0-10)",
        "start_at_zero": False,
        "pivot_for_chart": False,
        "data_type": "Number",
        "multilevel_headers": False,
        "source": {"Table ID": "", "URL": "https://www.countyhealthrankings.org/",
                   "Data Year": 2025, "Estimate Type": "",
                   "Citation Month": "April", "Citation Year": 2025,
                   "Custom Text": "County Health Rankings and Roadmaps, 2025"},
        "data": [
            {"County": "Dutchess", "Food Environment Index": 8.4},
            {"County": "Orange", "Food Environment Index": 8.3},
            {"County": "Putnam", "Food Environment Index": 9.0},
            {"County": "Rockland", "Food Environment Index": 8.5},
            {"County": "Sullivan", "Food Environment Index": 8.2},
            {"County": "Ulster", "Food Environment Index": 8.1},
            {"County": "Westchester", "Food Environment Index": 9.3},
            {"County": "NYS", "Food Environment Index": 8.7},
        ],
    },
    # 21 ── limited-access-foods
    {
        "sheet_name": "Limited-Access-Foods",
        "object_id": "limited-access-healthy-foods",
        "caption": "Percentage of Population with Limited Access to Healthy Foods, 2025",
        "type": "both",
        "figure_type": "Clustered Bar",
        "x_col": "County",
        "y_cols": "Limited Access",
        "y_axis_title": "Percentage with Limited Access (%)",
        "start_at_zero": False,
        "pivot_for_chart": False,
        "data_type": "Percent",
        "multilevel_headers": False,
        "source": {"Table ID": "", "URL": "https://www.countyhealthrankings.org/",
                   "Data Year": 2025, "Estimate Type": "",
                   "Citation Month": "April", "Citation Year": 2025,
                   "Custom Text": "County Health Rankings and Roadmaps, 2025"},
        "data": [
            {"County": "Dutchess", "Limited Access": 2.1},
            {"County": "Orange", "Limited Access": 2.8},
            {"County": "Putnam", "Limited Access": 6.7},
            {"County": "Rockland", "Limited Access": 2.5},
            {"County": "Sullivan", "Limited Access": 3.2},
            {"County": "Ulster", "Limited Access": 2.3},
            {"County": "Westchester", "Limited Access": 1.4},
            {"County": "NYS", "Limited Access": 2.0},
        ],
    },
    # 22 ── violent-crime
    {
        "sheet_name": "Violent-Crime",
        "object_id": "violent-crime",
        "caption": "Violent Crime Rate per 100,000 Population, 2018-2021",
        "type": "both",
        "figure_type": "Clustered Bar",
        "x_col": "Year",
        "y_cols": "",
        "y_axis_title": "Violent Crime Rate per 100,000",
        "start_at_zero": False,
        "pivot_for_chart": True,
        "data_type": "Number",
        "multilevel_headers": False,
        "source": {"Table ID": "", "URL": "https://www.criminaljustice.ny.gov/crimnet/ojsa/arrests/index.htm",
                   "Data Year": 2021, "Estimate Type": "",
                   "Citation Month": "April", "Citation Year": 2025,
                   "Custom Text": "NYS Division of Criminal Justice Services, April 2025"},
        "data": [
            {"Year": 2018, "Dutchess": 118.0, "Orange": 185.0, "Putnam": 45.0, "Rockland": 95.0,
             "Sullivan": 165.0, "Ulster": 125.0, "Westchester": 110.0, "NYS": 380.0},
            {"Year": 2019, "Dutchess": 120.0, "Orange": 190.0, "Putnam": 48.0, "Rockland": 98.0,
             "Sullivan": 170.0, "Ulster": 128.0, "Westchester": 112.0, "NYS": 375.0},
            {"Year": 2020, "Dutchess": 115.0, "Orange": 188.0, "Putnam": 42.0, "Rockland": 92.0,
             "Sullivan": 155.0, "Ulster": 120.0, "Westchester": 108.0, "NYS": 370.0},
            {"Year": 2021, "Dutchess": 112.0, "Orange": 192.4, "Putnam": 40.0, "Rockland": 88.0,
             "Sullivan": 145.0, "Ulster": 118.0, "Westchester": 105.0, "NYS": 365.0},
        ],
    },
    # 23 ── air-pollution
    {
        "sheet_name": "Air-Pollution",
        "object_id": "air-pollution",
        "caption": "Average Daily Density of Fine Particulate Matter, 2014-2020",
        "type": "both",
        "figure_type": "Line",
        "x_col": "Year",
        "y_cols": "",
        "y_axis_title": "Fine Particulate Matter (\u03bcg/m\u00b3)",
        "start_at_zero": False,
        "pivot_for_chart": False,
        "data_type": "Number",
        "multilevel_headers": False,
        "source": {"Table ID": "", "URL": "https://www.countyhealthrankings.org/",
                   "Data Year": 2020, "Estimate Type": "",
                   "Citation Month": "April", "Citation Year": 2025,
                   "Custom Text": "County Health Rankings and Roadmaps, 2025"},
        "data": [
            {"Year": 2014, "Dutchess": 8.5, "Orange": 7.8, "Putnam": 8.2, "Rockland": 8.8,
             "Sullivan": 7.5, "Ulster": 8.0, "Westchester": 9.2, "NYS": 8.5},
            {"Year": 2015, "Dutchess": 8.2, "Orange": 7.5, "Putnam": 8.0, "Rockland": 8.5,
             "Sullivan": 7.2, "Ulster": 7.8, "Westchester": 9.0, "NYS": 8.3},
            {"Year": 2016, "Dutchess": 7.9, "Orange": 7.2, "Putnam": 7.8, "Rockland": 8.2,
             "Sullivan": 7.0, "Ulster": 7.5, "Westchester": 8.8, "NYS": 8.1},
            {"Year": 2017, "Dutchess": 7.6, "Orange": 7.0, "Putnam": 7.5, "Rockland": 8.0,
             "Sullivan": 6.8, "Ulster": 7.3, "Westchester": 8.6, "NYS": 7.9},
            {"Year": 2018, "Dutchess": 7.4, "Orange": 6.8, "Putnam": 7.3, "Rockland": 7.8,
             "Sullivan": 6.6, "Ulster": 7.1, "Westchester": 8.4, "NYS": 7.7},
            {"Year": 2019, "Dutchess": 7.2, "Orange": 6.6, "Putnam": 7.1, "Rockland": 7.6,
             "Sullivan": 6.4, "Ulster": 6.9, "Westchester": 8.2, "NYS": 7.5},
            {"Year": 2020, "Dutchess": 7.0, "Orange": 6.4, "Putnam": 6.9, "Rockland": 7.4,
             "Sullivan": 6.2, "Ulster": 6.7, "Westchester": 8.0, "NYS": 7.3},
        ],
    },
    # 24 ── lead-testing
    {
        "sheet_name": "Lead-Testing",
        "object_id": "lead-testing",
        "caption": "Percentage of Children Tested for Lead at Least Twice Before 36 Months of Age, 2016-2019 Birth Cohorts",
        "type": "both",
        "figure_type": "Clustered Bar",
        "x_col": "Birth Cohort",
        "y_cols": "",
        "y_axis_title": "Percentage Tested (%)",
        "start_at_zero": False,
        "pivot_for_chart": True,
        "data_type": "Percent",
        "multilevel_headers": False,
        "source": {"Table ID": "", "URL": "https://webbi1.health.ny.gov/SASStoredProcess/guest?_program=%2FEBI%2FPHIG%2Fapps%2Fchir_dashboard%2Fchir_dashboard&p=ch&cos=33",
                   "Data Year": 2019, "Estimate Type": "",
                   "Citation Month": "April", "Citation Year": 2025,
                   "Custom Text": "NYS Community Health Indicator Reports Dashboard, April 2025"},
        "data": [
            {"Birth Cohort": 2016, "Dutchess": 45.2, "Orange": 42.3, "Putnam": 48.5,
             "Rockland": 44.8, "Sullivan": 32.5, "Ulster": 43.2, "Westchester": 56.8, "NYS": 52.5},
            {"Birth Cohort": 2017, "Dutchess": 46.8, "Orange": 43.7, "Putnam": 50.2,
             "Rockland": 46.2, "Sullivan": 34.1, "Ulster": 44.6, "Westchester": 58.2, "NYS": 53.8},
            {"Birth Cohort": 2018, "Dutchess": 48.1, "Orange": 45.2, "Putnam": 51.8,
             "Rockland": 47.5, "Sullivan": 35.2, "Ulster": 45.9, "Westchester": 59.5, "NYS": 55.2},
            {"Birth Cohort": 2019, "Dutchess": 49.5, "Orange": 46.8, "Putnam": 53.2,
             "Rockland": 48.9, "Sullivan": 36.6, "Ulster": 47.2, "Westchester": 60.2, "NYS": 56.5},
        ],
    },
    # 25 ── severe-housing-problems
    {
        "sheet_name": "Severe-Housing-Problems",
        "object_id": "severe-housing-problems",
        "caption": "Percentage of Households with Severe Housing Problems, 2025",
        "type": "both",
        "figure_type": "Clustered Bar",
        "x_col": "County",
        "y_cols": "Severe Housing Problems",
        "y_axis_title": "Percentage with Severe Housing Problems (%)",
        "start_at_zero": False,
        "pivot_for_chart": False,
        "data_type": "Percent",
        "multilevel_headers": False,
        "source": {"Table ID": "", "URL": "https://www.countyhealthrankings.org/",
                   "Data Year": 2025, "Estimate Type": "",
                   "Citation Month": "April", "Citation Year": 2025,
                   "Custom Text": "County Health Rankings and Roadmaps, 2025"},
        "data": [
            {"County": "Dutchess", "Severe Housing Problems": 20.0},
            {"County": "Orange", "Severe Housing Problems": 22.0},
            {"County": "Putnam", "Severe Housing Problems": 17.0},
            {"County": "Rockland", "Severe Housing Problems": 26.0},
            {"County": "Sullivan", "Severe Housing Problems": 16.0},
            {"County": "Ulster", "Severe Housing Problems": 19.0},
            {"County": "Westchester", "Severe Housing Problems": 21.0},
            {"County": "NYS", "Severe Housing Problems": 23.0},
        ],
    },
    # 26 ── no-vehicles
    {
        "sheet_name": "No-Vehicles",
        "object_id": "no-vehicles",
        "caption": "Percentage of Households with No Available Vehicles, 2023",
        "type": "both",
        "figure_type": "Clustered Bar",
        "x_col": "County",
        "y_cols": "No Vehicles",
        "y_axis_title": "Percentage with No Vehicles (%)",
        "start_at_zero": False,
        "pivot_for_chart": False,
        "data_type": "Percent",
        "multilevel_headers": False,
        "source": {"Table ID": "DP04", "URL": "https://data.census.gov/table/ACSDP5Y2023.DP04",
                   "Data Year": 2023, "Estimate Type": "5-Year Estimates",
                   "Citation Month": "April", "Citation Year": 2025, "Custom Text": ""},
        "data": [
            {"County": "Dutchess", "No Vehicles": 6.8},
            {"County": "Orange", "No Vehicles": 8.2},
            {"County": "Putnam", "No Vehicles": 4.2},
            {"County": "Rockland", "No Vehicles": 9.5},
            {"County": "Sullivan", "No Vehicles": 7.2},
            {"County": "Ulster", "No Vehicles": 6.5},
            {"County": "Westchester", "No Vehicles": 14.2},
            {"County": "NYS", "No Vehicles": 10.5},
        ],
    },
    # 27 ── modes-transportation
    {
        "sheet_name": "Modes-Transportation",
        "object_id": "modes-transportation",
        "caption": "Modes of Transportation to Work, 2023",
        "type": "both",
        "figure_type": "Stacked Bar",
        "x_col": "Mode",
        "y_cols": "",
        "y_axis_title": "Percentage (%)",
        "start_at_zero": True,
        "pivot_for_chart": False,
        "data_type": "Percent",
        "multilevel_headers": False,
        "source": {"Table ID": "S0801", "URL": "https://data.census.gov/table/ACSST5Y2023.S0801",
                   "Data Year": 2023, "Estimate Type": "5-Year Estimates",
                   "Citation Month": "April", "Citation Year": 2025, "Custom Text": ""},
        "data": [
            {"Mode": "Drive alone", "Dutchess": 68.5, "Orange": 70.2, "Putnam": 65.8,
             "Rockland": 66.5, "Sullivan": 74.1, "Ulster": 63.2, "Westchester": 51.9, "NYS": 58.5},
            {"Mode": "Carpool", "Dutchess": 8.2, "Orange": 7.8, "Putnam": 9.5,
             "Rockland": 8.8, "Sullivan": 6.5, "Ulster": 7.5, "Westchester": 6.8, "NYS": 7.2},
            {"Mode": "Public transportation", "Dutchess": 3.5, "Orange": 2.5, "Putnam": 4.2,
             "Rockland": 8.5, "Sullivan": 1.2, "Ulster": 2.8, "Westchester": 17.8, "NYS": 12.5},
            {"Mode": "Walk", "Dutchess": 2.8, "Orange": 2.2, "Putnam": 2.5,
             "Rockland": 2.8, "Sullivan": 1.8, "Ulster": 3.2, "Westchester": 3.5, "NYS": 3.8},
            {"Mode": "Other", "Dutchess": 1.2, "Orange": 1.1, "Putnam": 1.3,
             "Rockland": 1.2, "Sullivan": 0.9, "Ulster": 1.0, "Westchester": 1.2, "NYS": 1.5},
            {"Mode": "Work from home", "Dutchess": 12.8, "Orange": 11.2, "Putnam": 13.7,
             "Rockland": 8.2, "Sullivan": 8.5, "Ulster": 15.3, "Westchester": 17.3, "NYS": 12.5},
        ],
    },
    # 28 ── commute-time
    {
        "sheet_name": "Commute-Time",
        "object_id": "commute-time",
        "caption": "Average Commute Time to Work, 2023",
        "type": "both",
        "figure_type": "Clustered Bar",
        "x_col": "County",
        "y_cols": "Average Commute Time (minutes)",
        "y_axis_title": "Average Commute Time (minutes)",
        "start_at_zero": False,
        "pivot_for_chart": False,
        "data_type": "Number",
        "multilevel_headers": False,
        "source": {"Table ID": "S0801", "URL": "https://data.census.gov/table/ACSST5Y2023.S0801",
                   "Data Year": 2023, "Estimate Type": "5-Year Estimates",
                   "Citation Month": "April", "Citation Year": 2025, "Custom Text": ""},
        "data": [
            {"County": "Dutchess", "Average Commute Time (minutes)": 28},
            {"County": "Orange", "Average Commute Time (minutes)": 34},
            {"County": "Putnam", "Average Commute Time (minutes)": 39},
            {"County": "Rockland", "Average Commute Time (minutes)": 32},
            {"County": "Sullivan", "Average Commute Time (minutes)": 25},
            {"County": "Ulster", "Average Commute Time (minutes)": 25},
            {"County": "Westchester", "Average Commute Time (minutes)": 35},
            {"County": "NYS", "Average Commute Time (minutes)": 33},
        ],
    },
]


# ── helper: write one flat indicator sheet ────────────────────────────────────

def _nan_to_s(val):
    """Convert NaN/None to empty string; leave other values unchanged."""
    if val is None:
        return ""
    if isinstance(val, float) and math.isnan(val):
        return "s"
    return val


def _write_flat_sheet(wb: openpyxl.Workbook, indicator: dict) -> None:
    """
    Write a single flat indicator sheet in the format expected by
    workbook_loader._parse_flat_indicator_sheet().

    Layout (1-based row numbers, 1-based column numbers):
      Row 1:  A=Name,             B=caption
      Row 2:  A=Table/Figure/Both, B=Both|Table
      Row 3:  A=Object ID,        B=slug
      Row 4:  (empty)
      Row 5:  A=Table Rules,      E=Enter Data, F=h1, G=h2, ...
      Row 6:  A=Multilevel Headers, B=True/False, F=d[0][h1], G=d[0][h2], ...
      Row 7:  A=Data Type,        B=Percent|Number, F=d[1][h1], ...
      Row 8+: (no config key),    F=d[2][h1], ...  (extra data rows)
      Row 5+N+1: A=Figure Rules   (col F empty → stops data reading)
      Row 5+N+2: A=Figure Type,   B=type
      Row 5+N+3: A=X Column,      B=col
      Row 5+N+4: A=Y Column,      B=cols
      Row 5+N+5: A=X Axis Title   (unused by loader, visual label only)
      Row 5+N+6: A=Y Axis Title,  B=title
      Row 5+N+7: A=Start at Zero, B=True|False
      Row 5+N+8: A=Pivot For Chart, B=True|False
      Row 5+N+9: (empty)
      Row 5+N+10: A=Source Specifications
      Row 5+N+11: A=Table ID,     B=value
      Row 5+N+12: A=URL,          B=value
      Row 5+N+13: A=Data Year,    B=value
      Row 5+N+14: A=Estimate Type, B=value
      Row 5+N+15: A=Citation Month, B=value
      Row 5+N+16: A=Citation Year, B=value
      Row 5+N+17: A=Custom Text,  B=value
    """
    sheet_name = indicator["sheet_name"]

    # Skip if sheet already exists (idempotent)
    if sheet_name in wb.sheetnames:
        print(f"  [skip] Sheet '{sheet_name}' already exists.")
        return

    ws = wb.create_sheet(title=sheet_name)

    data = indicator["data"]
    if not data:
        return

    headers = list(data[0].keys())
    n_data_rows = len(data)

    is_table_only = indicator["type"] == "table"
    type_val = "Table" if is_table_only else "Both"
    pivot_val = "True" if indicator["pivot_for_chart"] else "False"
    start_zero_val = "True" if indicator["start_at_zero"] else "False"
    multilevel_val = "True" if indicator["multilevel_headers"] else "False"

    # ── fixed config rows (rows 1-4) ──────────────────────────────────────────
    ws.cell(row=1, column=COL_KEY, value="Name")
    ws.cell(row=1, column=COL_VAL, value=indicator["caption"])
    ws.cell(row=2, column=COL_KEY, value="Table/Figure/Both")
    ws.cell(row=2, column=COL_VAL, value=type_val)
    ws.cell(row=3, column=COL_KEY, value="Object ID")
    ws.cell(row=3, column=COL_VAL, value=indicator["object_id"])
    # Row 4 is empty

    # ── table rules + data section (row 5 = header row) ───────────────────────
    HEADER_ROW = 5
    ws.cell(row=HEADER_ROW, column=COL_KEY, value=SEC_TABLE_RULES)
    ws.cell(row=HEADER_ROW, column=COL_MARKER, value="Enter Data")
    for col_offset, col_name in enumerate(headers):
        ws.cell(row=HEADER_ROW, column=COL_DATA_START + col_offset, value=col_name)

    # Inline table config keys alongside data rows
    table_config_keys = [
        ("Multilevel Headers", multilevel_val),
        ("Data Type", indicator["data_type"]),
    ]

    for row_offset, row_dict in enumerate(data):
        r = HEADER_ROW + 1 + row_offset
        # Write config key/value if we have one for this row index
        if row_offset < len(table_config_keys):
            key, val = table_config_keys[row_offset]
            ws.cell(row=r, column=COL_KEY, value=key)
            ws.cell(row=r, column=COL_VAL, value=val)
        # Write data values in cols F+
        for col_offset, col_name in enumerate(headers):
            raw_val = _nan_to_s(row_dict.get(col_name, ""))
            ws.cell(row=r, column=COL_DATA_START + col_offset, value=raw_val)

    # The row immediately after the last data row must have empty col F to stop
    # data reading. "Figure Rules" section header serves this purpose.
    FIGURE_RULES_ROW = HEADER_ROW + 1 + n_data_rows  # col F empty here → stops data

    # ── figure rules section ──────────────────────────────────────────────────
    ws.cell(row=FIGURE_RULES_ROW, column=COL_KEY, value=SEC_FIGURE_RULES)
    if not is_table_only:
        ws.cell(row=FIGURE_RULES_ROW + 1, column=COL_KEY, value="Figure Type")
        ws.cell(row=FIGURE_RULES_ROW + 1, column=COL_VAL, value=indicator["figure_type"])
        ws.cell(row=FIGURE_RULES_ROW + 2, column=COL_KEY, value="X Column")
        ws.cell(row=FIGURE_RULES_ROW + 2, column=COL_VAL, value=indicator["x_col"])
        ws.cell(row=FIGURE_RULES_ROW + 3, column=COL_KEY, value="Y Column")
        ws.cell(row=FIGURE_RULES_ROW + 3, column=COL_VAL, value=indicator["y_cols"])
        ws.cell(row=FIGURE_RULES_ROW + 4, column=COL_KEY, value="Y Axis Title")
        ws.cell(row=FIGURE_RULES_ROW + 4, column=COL_VAL, value=indicator["y_axis_title"])
        ws.cell(row=FIGURE_RULES_ROW + 5, column=COL_KEY, value="Start at Zero")
        ws.cell(row=FIGURE_RULES_ROW + 5, column=COL_VAL, value=start_zero_val)
        ws.cell(row=FIGURE_RULES_ROW + 6, column=COL_KEY, value="Pivot For Chart")
        ws.cell(row=FIGURE_RULES_ROW + 6, column=COL_VAL, value=pivot_val)
        src_base = FIGURE_RULES_ROW + 8
    else:
        src_base = FIGURE_RULES_ROW + 2

    # ── source specifications section ─────────────────────────────────────────
    src = indicator.get("source", {})
    ws.cell(row=src_base, column=COL_KEY, value=SEC_SOURCE_SPECS)
    ws.cell(row=src_base + 1, column=COL_KEY, value="Table ID")
    ws.cell(row=src_base + 1, column=COL_VAL, value=src.get("Table ID", ""))
    ws.cell(row=src_base + 2, column=COL_KEY, value="URL")
    ws.cell(row=src_base + 2, column=COL_VAL, value=src.get("URL", ""))
    ws.cell(row=src_base + 3, column=COL_KEY, value="Data Year")
    ws.cell(row=src_base + 3, column=COL_VAL, value=src.get("Data Year", ""))
    ws.cell(row=src_base + 4, column=COL_KEY, value="Estimate Type")
    ws.cell(row=src_base + 4, column=COL_VAL, value=src.get("Estimate Type", ""))
    ws.cell(row=src_base + 5, column=COL_KEY, value="Citation Month")
    ws.cell(row=src_base + 5, column=COL_VAL, value=src.get("Citation Month", ""))
    ws.cell(row=src_base + 6, column=COL_KEY, value="Citation Year")
    ws.cell(row=src_base + 6, column=COL_VAL, value=src.get("Citation Year", ""))
    ws.cell(row=src_base + 7, column=COL_KEY, value="Custom Text")
    ws.cell(row=src_base + 7, column=COL_VAL, value=src.get("Custom Text", ""))


def _fill_source_specs_on_existing_sheet(wb: openpyxl.Workbook, sheet_name: str, specs: dict) -> None:
    """
    Find the Source Specifications section on an existing sheet and fill in
    blank cells.  Adds new rows if the spec keys are missing.
    """
    if sheet_name not in wb.sheetnames:
        print(f"  [warn] Sheet '{sheet_name}' not found; cannot fill source specs.")
        return

    ws = wb[sheet_name]
    SOURCE_KEYS = ["Table ID", "URL", "Data Year", "Estimate Type",
                   "Citation Month", "Citation Year", "Custom Text"]

    # Find existing source spec rows
    src_section_row = None
    key_rows: dict[str, int] = {}
    for row in ws.iter_rows():
        cell_a = row[0].value
        if cell_a == SEC_SOURCE_SPECS:
            src_section_row = row[0].row
        elif cell_a in SOURCE_KEYS:
            key_rows[cell_a] = row[0].row

    if src_section_row is None:
        print(f"  [warn] No 'Source Specifications' section found in '{sheet_name}'.")
        return

    # Fill or add each key
    for key in SOURCE_KEYS:
        value = specs.get(key, "")
        if key in key_rows:
            r = key_rows[key]
            existing_b = ws.cell(row=r, column=COL_VAL).value
            if existing_b is None or existing_b == "":
                ws.cell(row=r, column=COL_VAL, value=value)
        else:
            # Append after the last known source row or after section header
            last_row = max(key_rows.values()) if key_rows else src_section_row
            new_row = last_row + 1
            ws.cell(row=new_row, column=COL_KEY, value=key)
            ws.cell(row=new_row, column=COL_VAL, value=value)
            key_rows[key] = new_row

    print(f"  [ok] Filled source specs on '{sheet_name}'.")


# ── main ──────────────────────────────────────────────────────────────────────

def main() -> None:
    if not WORKBOOK_PATH.exists():
        raise FileNotFoundError(f"Workbook not found: {WORKBOOK_PATH}")

    print(f"Loading workbook: {WORKBOOK_PATH.name}")
    wb = load_workbook(WORKBOOK_PATH)
    print(f"  Existing sheets: {wb.sheetnames}")

    # 1. Rename Template → _Template
    if "Template" in wb.sheetnames:
        wb["Template"].title = "_Template"
        print("  [ok] Renamed 'Template' → '_Template'")
    else:
        print("  [skip] 'Template' sheet not found (may already be renamed).")

    # 2. Remove dead sheets
    for dead in ("data_fig_unemployment", "data_tbl_unemployment"):
        if dead in wb.sheetnames:
            del wb[dead]
            print(f"  [ok] Deleted sheet '{dead}'")
        else:
            print(f"  [skip] Sheet '{dead}' not found.")

    # 3. Fill source specs for existing sheets
    print("\nFilling source specs for existing sheets...")
    for sheet_name, specs in EXISTING_SOURCE_SPECS.items():
        _fill_source_specs_on_existing_sheet(wb, sheet_name, specs)

    # 4. Add new indicator sheets (skip if sheet already exists)
    print("\nAdding ch04 indicator sheets...")
    for indicator in CH04_INDICATORS:
        _write_flat_sheet(wb, indicator)
        status = "already exists" if indicator["sheet_name"] in wb.sheetnames else "added"
        print(f"  {indicator['sheet_name']}: {status}")

    # Save
    wb.save(WORKBOOK_PATH)
    print(f"\nSaved: {WORKBOOK_PATH}")
    print(f"Sheets now: {wb.sheetnames}")


if __name__ == "__main__":
    main()
